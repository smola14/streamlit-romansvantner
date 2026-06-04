from __future__ import annotations

import hmac
import hashlib
import base64
import io
import json
import os
import struct
import sys
import unicodedata
from collections.abc import Mapping
from pathlib import Path
from datetime import date, datetime, time, timedelta, timezone
from typing import Any

import matplotlib.pyplot as plt
from matplotlib import font_manager
from openpyxl import load_workbook
import requests
import streamlit as st
from fpdf import FPDF

APP_DIR = Path(__file__).resolve().parent
if str(APP_DIR) not in sys.path:
    sys.path.insert(0, str(APP_DIR))

from report_deceleration import (
    build_non_normative_deceleration_pdf as decel_build_non_normative_deceleration_pdf,
    make_deceleration_speed_time_plot as decel_make_deceleration_speed_time_plot,
)
from report_fv import (
    build_non_normative_fv_pdf as fv_build_non_normative_fv_pdf,
    build_normative_fv_pdf as fv_build_normative_fv_pdf,
)
from report_split import build_non_normative_split_pdf as split_build_non_normative_split_pdf


API_BASE_URL = os.getenv("API1080_BASE_URL", "https://publicapi.1080motion.com").rstrip("/")
APP_API_KEY = st.secrets.get("api_1080_key", os.getenv("API1080_KEY", "")).strip()
AUTH_SESSION_SECRET = st.secrets.get("auth_session_secret", os.getenv("AUTH_SESSION_SECRET", "")).strip()
API_TIMEOUT_SECONDS = 20
BLUE_RGB = (48, 54, 116)
BLACK_RGB = (37, 36, 35)
GREEN_RGB = (0, 192, 96)
ORANGE_RGB = (254, 148, 65)
RED_RGB = (251, 51, 49)
BLUE_HEX = "#303674"
ORANGE_HEX = "#FE9441"
RED_HEX = "#FB3331"
FV_NORMS_PATH = Path(__file__).resolve().parent / "data" / "fv_norms.xlsx"
FV_NORM_SCATTER_PATH = Path(__file__).resolve().parent / "data" / "fv_norm_scatter.json"
RS_LOGO_PATH = Path(__file__).resolve().parent / "rs-logo.png"
SPLIT_1505_IMAGE_PATH = Path(__file__).resolve().parent / "1505.png"
DECEL_V_STOP = 0.2
DECEL_ACC_THRESHOLD = -1.5
DECEL_SAMPLE_FILTER_MODE = "Smooth"
UPLOADED_LOGOS_DIR = Path(__file__).resolve().parent / "uploaded_logos"
UPLOADED_PLAYER_PHOTOS_DIR = Path(__file__).resolve().parent / "uploaded_player_photos"
MAX_LOGO_BYTES = 2 * 1024 * 1024
MAX_PLAYER_PHOTO_BYTES = 5 * 1024 * 1024
AUTH_SESSION_DURATION = timedelta(days=1)
AUTH_STORAGE_KEY = "1080_auth_session"
RECENT_OPENED_SESSION_DURATION = timedelta(days=1)


PDF_TEXT = {
    "English": {
        "fv_title": "Force-Velocity Profile",
        "decel_title": "Deceleration Profile",
        "decel_chart_title": "Deceleration trace",
        "decel_open_export": "Open deceleration PDF export",
        "decel_download_pdf": "Download deceleration PDF",
        "decel_choose_run": "Choose the deceleration run and optional assets for the export.",
        "decel_run": "Run",
        "decel_preview_run": "Preview run",
        "decel_runs": "Runs",
        "decel_threshold": "Threshold",
        "decel_stop_speed": "Stop speed",
        "decel_avg": "DecA [m/s²]",
        "decel_max": "DecM [m/s²]",
        "decel_vmax": "VMax [m/s]",
        "decel_tts": "TTS [s]",
        "decel_dts": "DTS [m]",
        "decel_no_runs": "No deceleration runs are available for PDF export.",
        "decel_no_valid_runs": "No valid deceleration runs were derived from the available training data.",
        "language": "Language",
        "slovak": "Slovak",
        "english": "English",
        "branding": "Branding",
        "choose_logo": "Choose logo",
        "no_logo_selected": "No logo selected",
        "selected_logo": "Selected logo",
        "clear_selection": "Clear selection",
        "upload_new_logo": "Upload new logo",
        "player_photo": "Player photo",
        "add_player_photo": "Add a player photo once and it will be reused for future exports.",
        "upload_player_photo": "Upload player photo",
        "early_dec": "Early dec",
        "late_dec": "Late dec",
        "recommendation": "Recommendation",
        "player_fv_profile": "Player FV profile",
        "quadrant_reference": "Quadrant reference",
        "upper_reference": "Upper reference",
        "player_label": "Player",
        "lower_reference": "Lower reference",
        "notes": "Notes",
        "f0_status": "F0 vs norm",
        "v0_status": "V0 vs norm",
        "within": "Within",
        "below": "Below",
        "above": "Above",
        "no_norm": "No norm",
        "q1_result": "Faster acceleration / higher speed",
        "q2_result": "Faster acceleration / lower speed",
        "q3_result": "Slower acceleration / lower speed",
        "q4_result": "Slower acceleration / higher speed",
        "q1_recs": [
            "Acceleration and speed work (<7 s)",
            "Resisted sprint training (25-50% speed drop; 10-20 m)",
            "Flying sprints",
            "Assisted sprint training",
            "Speed bounding",
            "Improve hamstring isometric strength",
        ],
        "q2_recs": [
            "Acceleration and speed work (<7 s)",
            "Resisted sprint training (25-50% speed drop; 10-20 m)",
            "Flying sprints",
            "Assisted sprint training",
            "Improve stretch-shortening cycle (SSC)",
            "Improve reactive strength",
            "Improve connective tissue strength",
        ],
        "q3_recs": [
            "Resisted sprint training (50-75% speed drop; 10 m)",
            "Flying sprints",
            "Improve stretch-shortening cycle (SSC)",
            "Improve hip extensor strength",
            "Improve soleus and gastrocnemius strength",
            "Improve absolute and relative strength",
            "Improve reactive strength",
            "Improve rate of force development (RFD)",
        ],
        "q4_recs": [
            "Acceleration and speed work (<7 s)",
            "Resisted sprint training (50-75% speed drop; 10 m)",
            "Improve hip extensor strength",
            "Improve absolute and relative strength",
            "Improve rate of force development (RFD)",
            "Improve connective tissue strength",
        ],
    },
    "Slovak": {
        "fv_title": "Silovo-rýchlostný profil",
        "decel_title": "Deceleračný profil",
        "decel_chart_title": "Priebeh decelerácie",
        "decel_open_export": "Otvoriť PDF export decelerácie",
        "decel_download_pdf": "Stiahnuť PDF decelerácie",
        "decel_choose_run": "Vyber pokus a voliteľné podklady pre export.",
        "decel_run": "Pokus",
        "decel_preview_run": "Náhľad pokusu",
        "decel_runs": "Pokusy",
        "decel_threshold": "Prahová hodnota",
        "decel_stop_speed": "Koncová rýchlosť",
        "decel_avg": "DecA [m/s²]",
        "decel_max": "DecM [m/s²]",
        "decel_vmax": "VMax [m/s]",
        "decel_tts": "TTS [s]",
        "decel_dts": "DTS [m]",
        "decel_no_runs": "Pre export nie sú dostupné žiadne deceleračné pokusy.",
        "decel_no_valid_runs": "Z dostupných tréningových dát sa nepodarilo odvodiť validné deceleračné pokusy.",
        "language": "Jazyk",
        "slovak": "Slovensky",
        "english": "Anglicky",
        "branding": "Branding",
        "choose_logo": "Vybrať logo",
        "no_logo_selected": "Nie je vybrané logo",
        "selected_logo": "Vybrané logo",
        "clear_selection": "Zrušiť výber",
        "upload_new_logo": "Nahrať nové logo",
        "player_photo": "Fotka hráča",
        "add_player_photo": "Fotku hráča stačí nahrať raz a použije sa aj pri ďalších exportoch.",
        "upload_player_photo": "Nahrať fotku hráča",
        "early_dec": "Skorá dec.",
        "late_dec": "Neskorá dec.",
        "recommendation": "Odporúčanie",
        "player_fv_profile": "Hráčsky FV profil",
        "quadrant_reference": "Kvadrantová referencia",
        "upper_reference": "Horná referencia",
        "player_label": "Hráč",
        "lower_reference": "Dolná referencia",
        "notes": "Poznámka",
        "f0_status": "F0 vs norma",
        "v0_status": "V0 vs norma",
        "within": "V norme",
        "below": "Pod normou",
        "above": "Nad normou",
        "no_norm": "Bez normy",
        "q1_result": "Rýchlejšia akcelerácia / vyššia rýchlosť",
        "q2_result": "Rýchlejšia akcelerácia / nižšia rýchlosť",
        "q3_result": "Pomalšia akcelerácia / nižšia rýchlosť",
        "q4_result": "Pomalšia akcelerácia / vyššia rýchlosť",
        "q1_recs": [
            "Akcelerácia/práca na rýchlosti (<7 s)",
            "Šprintérsky tréning s odporom (25-50% pokles rýchlosti; 10-20 m)",
            "Letmé šprinty",
            "Šprintérsky tréning s asistenciou",
            "Speed bounding",
            "Zlepšenie izometrickej sily hamstringov",
        ],
        "q2_recs": [
            "Akcelerácia/práca na rýchlosti (<7 s)",
            "Šprintérsky tréning s odporom (25-50% pokles rýchlosti; 10-20 m)",
            "Letmé šprinty",
            "Šprintérsky tréning s asistenciou",
            "Zlepšenie cyklu natiahnutie-skrátenie (SSC)",
            "Zlepšenie reaktívnej sily",
            "Zlepšenie sily spojivových tkanív",
        ],
        "q3_recs": [
            "Šprintérsky tréning s odporom (50-75% pokles rýchlosti; 10 m)",
            "Letmé šprinty",
            "Zlepšenie cyklu natiahnutia-skrátenia (SSC)",
            "Zlepšenie sily extenzorov bedrového kĺbu",
            "Zlepšenie sily soleusu a gastrocnemiusu",
            "Zlepšenie absolútnej/relatívnej sily",
            "Zlepšenie reaktívnej sily",
            "Zlepšenie rýchlosti produkcie sily (RFD)",
        ],
        "q4_recs": [
            "Akcelerácia/práca na rýchlosti (<7 s)",
            "Šprintérsky tréning s odporom (50-75% pokles rýchlosti; 10 m)",
            "Zlepšenie sily extenzorov bedrového kĺbu",
            "Zlepšenie absolútnej/relatívnej sily",
            "Zlepšenie rýchlosti produkcie sily (RFD)",
            "Zlepšenie sily spojivových tkanív",
        ],
    },
}

CLIENT_STORAGE_COMPONENT = st.components.v2.component(
    "client_storage",
    html="""
    <div id="client-storage-root"></div>
    """,
    js="""
    export default function(component) {
      const { data, setStateValue } = component;
      const storageKey = data?.storageKey ?? "1080_clients";
      const syncedKey = `${storageKey}:last_synced`;
      const currentCommandId = data?.commandId ?? "";
      const command = data?.command ?? "read";
      const recentOpenedStorageKey = data?.recentOpenedStorageKey ?? "1080_recent_opened_sessions";
      const recentOpenedCommand = data?.recentOpenedCommand ?? "read";
      const recentOpenedCommandId = data?.recentOpenedCommandId ?? "";
      const authStorageKey = data?.authStorageKey ?? "1080_auth_session";
      const authCommand = data?.authCommand ?? "read";
      const authCommandId = data?.authCommandId ?? "";

      const emitState = () => {
        setStateValue("clients_json", localStorage.getItem(storageKey) ?? "");
        setStateValue("last_synced", localStorage.getItem(syncedKey) ?? "");
        setStateValue("recent_opened_sessions_json", localStorage.getItem(recentOpenedStorageKey) ?? "");
        setStateValue("auth_session_token", localStorage.getItem(authStorageKey) ?? "");
        setStateValue("auth_storage_ready", true);
      };

      const commandMarkerKey = `__last_command__:${storageKey}`;
      const lastCommandId = window[commandMarkerKey];

      if (currentCommandId && currentCommandId !== lastCommandId) {
        if (command === "write") {
          localStorage.setItem(storageKey, data?.clientsJson ?? "");
          localStorage.setItem(syncedKey, data?.lastSynced ?? "");
        } else if (command === "clear") {
          localStorage.removeItem(storageKey);
          localStorage.removeItem(syncedKey);
        }

        window[commandMarkerKey] = currentCommandId;
      }

      const recentOpenedCommandMarkerKey = `__last_recent_opened_command__:${recentOpenedStorageKey}`;
      const lastRecentOpenedCommandId = window[recentOpenedCommandMarkerKey];

      if (recentOpenedCommandId && recentOpenedCommandId !== lastRecentOpenedCommandId) {
        if (recentOpenedCommand === "write") {
          localStorage.setItem(recentOpenedStorageKey, data?.recentOpenedSessionsJson ?? "");
        } else if (recentOpenedCommand === "clear") {
          localStorage.removeItem(recentOpenedStorageKey);
        }

        window[recentOpenedCommandMarkerKey] = recentOpenedCommandId;
      }

      const authCommandMarkerKey = `__last_auth_command__:${authStorageKey}`;
      const lastAuthCommandId = window[authCommandMarkerKey];

      if (authCommandId && authCommandId !== lastAuthCommandId) {
        if (authCommand === "write") {
          localStorage.setItem(authStorageKey, data?.authSessionToken ?? "");
        } else if (authCommand === "clear") {
          localStorage.removeItem(authStorageKey);
        }

        window[authCommandMarkerKey] = authCommandId;
      }

      emitState();
    }
    """,
)


def auth_session_signature(payload: str) -> str:
    if not AUTH_SESSION_SECRET:
        return ""
    return hmac.new(
        AUTH_SESSION_SECRET.encode("utf-8"),
        payload.encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()


def create_auth_session_token(email: str) -> str:
    expires_at = int((datetime.now(timezone.utc) + AUTH_SESSION_DURATION).timestamp())
    payload = json.dumps({"email": email, "exp": expires_at}, separators=(",", ":"))
    payload_b64 = base64.urlsafe_b64encode(payload.encode("utf-8")).decode("ascii")
    signature = auth_session_signature(payload_b64)
    return f"{payload_b64}.{signature}"


def parse_auth_session_token(token: str) -> str | None:
    if not token or not AUTH_SESSION_SECRET or "." not in token:
        return None

    payload_b64, provided_signature = token.rsplit(".", 1)
    expected_signature = auth_session_signature(payload_b64)
    if not expected_signature or not hmac.compare_digest(expected_signature, provided_signature):
        return None

    try:
        payload_json = base64.urlsafe_b64decode(payload_b64.encode("ascii")).decode("utf-8")
        payload = json.loads(payload_json)
        email = str(payload["email"]).strip().lower()
        expires_at = int(payload["exp"])
    except (ValueError, KeyError, TypeError, json.JSONDecodeError):
        return None

    if not email or datetime.now(timezone.utc).timestamp() > expires_at:
        return None

    return email


def queue_auth_storage_command(command: str, token: str = "") -> None:
    st.session_state["auth_storage_command"] = command
    st.session_state["auth_storage_command_id"] = f"{command}:{iso_now()}"
    st.session_state["auth_session_token"] = token


def restore_auth_session_from_token(token: str) -> bool:
    auth_users = get_auth_users()
    email = parse_auth_session_token(token)
    if not email or email not in auth_users or not APP_API_KEY:
        return False

    st.session_state["auth_verified"] = True
    st.session_state["auth_user_email"] = email
    st.session_state["api_key"] = APP_API_KEY
    st.session_state["api_valid"] = True
    st.session_state["client_storage_autoload_complete"] = False
    st.session_state["auth_storage_autoload_complete"] = True
    return True


def render_app_styles() -> None:
    st.markdown(
        """
        <style>
        .stApp {
            background: radial-gradient(circle at top, rgba(48,54,116,0.08), transparent 28%), #0f1117;
        }
        .block-container {
            padding-top: 2rem;
            padding-bottom: 3rem;
            max-width: 1240px;
        }
        .hero-card {
            padding: 1.25rem 1.25rem 0.25rem;
            border: 1px solid rgba(255,255,255,0.08);
            border-radius: 18px;
            background: linear-gradient(180deg, rgba(255,255,255,0.03), rgba(255,255,255,0.015));
            margin-bottom: 1rem;
        }
        .section-intro {
            color: rgba(255,255,255,0.72);
            margin-top: -0.35rem;
            margin-bottom: 0;
            font-size: 0.95rem;
        }
        div[data-testid="stMetric"] {
            border: 1px solid rgba(255,255,255,0.08);
            border-radius: 16px;
            padding: 0.9rem 1rem;
            background: rgba(255,255,255,0.02);
        }
        div[data-testid="stDataFrame"] {
            border-radius: 14px;
            overflow: hidden;
        }
        div[data-testid="stForm"] {
            border: 1px solid rgba(255,255,255,0.08);
            border-radius: 18px;
            padding: 1rem 1rem 0.25rem;
            background: rgba(255,255,255,0.02);
        }
        .thumb-card {
            border: 1px solid rgba(255,255,255,0.08);
            border-radius: 16px;
            padding: 0.75rem;
            background: rgba(255,255,255,0.02);
            width: fit-content;
        }
        .session-row {
            border: 1px solid rgba(255,255,255,0.08);
            border-radius: 16px;
            padding: 0.85rem 1rem;
            background: rgba(255,255,255,0.02);
            margin-bottom: 0.65rem;
        }
        .session-row-title {
            font-size: 1rem;
            font-weight: 600;
            margin: 0;
        }
        .session-row-subtitle {
            color: rgba(255,255,255,0.68);
            font-size: 0.92rem;
            margin: 0.18rem 0 0;
        }
        .session-row-inline {
            display: flex;
            align-items: baseline;
            gap: 0.65rem;
            flex-wrap: wrap;
        }
        .session-row-shell {
            display: flex;
            align-items: center;
            justify-content: space-between;
            gap: 1rem;
        }
        div[data-testid="stButton"] > button,
        div[data-testid="stDownloadButton"] > button,
        div[data-testid="stFormSubmitButton"] > button {
            background: #2f6df6;
            color: white;
            border: 1px solid #2f6df6;
        }
        div[data-testid="stButton"] > button:hover,
        div[data-testid="stDownloadButton"] > button:hover,
        div[data-testid="stFormSubmitButton"] > button:hover {
            background: #2559c9;
            border-color: #2559c9;
            color: white;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


def build_headers(api_key: str) -> dict[str, str]:
    return {
        "X-1080-API-Key": api_key,
        "Accept": "application/json",
    }


def validate_api_key(api_key: str) -> tuple[bool, str]:
    try:
        response = requests.get(
            f"{API_BASE_URL}/Session",
            headers=build_headers(api_key),
            params={"maxAgeDays": 0},
            timeout=API_TIMEOUT_SECONDS,
        )
    except requests.RequestException as exc:
        return False, f"Communication error while connecting to the 1080 API: {exc}"

    if response.status_code == 200:
        return True, "API key is valid."

    if response.status_code in (401, 403):
        return False, "API key is invalid or does not have access to the requested data."

    return False, f"1080 API returned an unexpected response: {response.status_code}"


def fetch_clients(api_key: str) -> tuple[list[dict[str, Any]] | None, str | None]:
    try:
        response = requests.get(
            f"{API_BASE_URL}/Client",
            headers=build_headers(api_key),
            timeout=API_TIMEOUT_SECONDS,
        )
    except requests.RequestException as exc:
        return None, f"Communication error while loading clients: {exc}"

    if response.status_code == 200:
        payload = response.json()
        if isinstance(payload, list):
            return payload, None
        return None, "The client response did not return the expected list format."

    if response.status_code in (401, 403):
        return None, "The API key is no longer authorized to load clients."

    return None, f"1080 API returned an unexpected response while loading clients: {response.status_code}"


def fetch_client_sessions(
    api_key: str,
    client_id: str,
    from_date: date,
    to_date: date,
) -> tuple[list[dict[str, Any]] | None, str | None]:
    from_dt = datetime.combine(from_date, time.min, tzinfo=timezone.utc)
    to_dt = datetime.combine(to_date, time.max, tzinfo=timezone.utc)

    try:
        response = requests.get(
            f"{API_BASE_URL}/Session/Search",
            headers=build_headers(api_key),
            params={
                "client": client_id,
                "fromDate": from_dt.isoformat(),
                "toDate": to_dt.isoformat(),
            },
            timeout=API_TIMEOUT_SECONDS,
        )
    except requests.RequestException as exc:
        return None, f"Communication error while loading sessions: {exc}"

    if response.status_code == 200:
        payload = response.json()
        if isinstance(payload, list):
            return payload, None
        return None, "The session response did not return the expected list format."

    if response.status_code in (401, 403):
        return None, "The API key is no longer authorized to load sessions."

    return None, f"1080 API returned an unexpected response while loading sessions: {response.status_code}"


def fetch_recent_sessions(
    api_key: str,
    from_date: date,
    to_date: date,
) -> tuple[list[dict[str, Any]] | None, str | None]:
    from_dt = datetime.combine(from_date, time.min, tzinfo=timezone.utc)
    to_dt = datetime.combine(to_date, time.max, tzinfo=timezone.utc)

    try:
        response = requests.get(
            f"{API_BASE_URL}/Session/Search",
            headers=build_headers(api_key),
            params={
                "fromDate": from_dt.isoformat(),
                "toDate": to_dt.isoformat(),
            },
            timeout=API_TIMEOUT_SECONDS,
        )
    except requests.RequestException as exc:
        return None, f"Communication error while loading recent sessions: {exc}"

    if response.status_code == 200:
        payload = response.json()
        if isinstance(payload, list):
            return payload, None
        return None, "The recent sessions response did not return the expected list format."

    if response.status_code in (401, 403):
        return None, "The API key is no longer authorized to load recent sessions."

    return None, f"1080 API returned an unexpected response while loading recent sessions: {response.status_code}"


def fetch_session_detail(api_key: str, session_id: str) -> tuple[dict[str, Any] | None, str | None]:
    try:
        response = requests.get(
            f"{API_BASE_URL}/Session/{session_id}",
            headers=build_headers(api_key),
            timeout=API_TIMEOUT_SECONDS,
        )
    except requests.RequestException as exc:
        return None, f"Communication error while loading session detail: {exc}"

    if response.status_code == 200:
        payload = response.json()
        if isinstance(payload, dict):
            return payload, None
        return None, "The session detail response did not return the expected object format."

    if response.status_code in (401, 403):
        return None, "The API key is no longer authorized to load session detail."

    return None, f"1080 API returned an unexpected response while loading session detail: {response.status_code}"


def fetch_force_velocity_exercise(
    api_key: str, exercise_id: str
) -> tuple[dict[str, Any] | None, str | None]:
    try:
        response = requests.get(
            f"{API_BASE_URL}/ForceVelocity/Exercise/{exercise_id}",
            headers=build_headers(api_key),
            timeout=API_TIMEOUT_SECONDS,
        )
    except requests.RequestException as exc:
        return None, f"Communication error while loading FV profile: {exc}"

    if response.status_code == 200:
        payload = response.json()
        if isinstance(payload, dict):
            return payload, None
        return None, "The FV response did not return the expected object format."

    if response.status_code in (401, 403):
        return None, "The API key is no longer authorized to load FV profile data."

    return None, f"1080 API returned an unexpected response while loading FV profile: {response.status_code}"


def fetch_split_exercise(
    api_key: str, exercise_id: str
) -> tuple[dict[str, Any] | None, str | None]:
    try:
        response = requests.get(
            f"{API_BASE_URL}/Split/Exercise/{exercise_id}",
            headers=build_headers(api_key),
            params={
                "splitLength": 5,
                "useYards": False,
                "includeRawPeaksAndAverages": True,
            },
            timeout=API_TIMEOUT_SECONDS,
        )
    except requests.RequestException as exc:
        return None, f"Communication error while loading split data: {exc}"

    if response.status_code == 200:
        payload = response.json()
        if isinstance(payload, dict):
            return payload, None
        return None, "The split response did not return the expected object format."

    if response.status_code in (401, 403):
        return None, "The API key is no longer authorized to load split data."

    return None, f"1080 API returned an unexpected response while loading split data: {response.status_code}"


def fetch_split_set(
    api_key: str, set_id: str
) -> tuple[dict[str, Any] | None, str | None]:
    try:
        response = requests.get(
            f"{API_BASE_URL}/Split/Set/{set_id}",
            headers=build_headers(api_key),
            params={
                "splitLength": 5,
                "useYards": False,
                "includeRawPeaksAndAverages": True,
            },
            timeout=API_TIMEOUT_SECONDS,
        )
    except requests.RequestException as exc:
        return None, f"Communication error while loading split set data: {exc}"

    if response.status_code == 200:
        payload = response.json()
        if isinstance(payload, dict):
            return payload, None
        return None, "The split set response did not return the expected object format."

    if response.status_code in (401, 403):
        return None, "The API key is no longer authorized to load split set data."

    return None, f"1080 API returned an unexpected response while loading split set data: {response.status_code}"


def fetch_training_data_set(
    api_key: str, set_id: str
) -> tuple[dict[str, Any] | None, str | None]:
    try:
        response = requests.get(
            f"{API_BASE_URL}/TrainingData/Set/{set_id}",
            headers=build_headers(api_key),
            params={
                "includeSamples": True,
                "filterMode": DECEL_SAMPLE_FILTER_MODE,
            },
            timeout=API_TIMEOUT_SECONDS,
        )
    except requests.RequestException as exc:
        return None, f"Communication error while loading set training data: {exc}"

    if response.status_code == 200:
        payload = response.json()
        if isinstance(payload, dict):
            return payload, None
        return None, "The set training data response did not return the expected object format."

    if response.status_code in (401, 403):
        return None, "The API key is no longer authorized to load set training data."

    return None, f"1080 API returned an unexpected response while loading set training data: {response.status_code}"


def fetch_split_runs(
    api_key: str, run_ids: list[str]
) -> tuple[list[dict[str, Any]] | None, str | None]:
    try:
        response = requests.get(
            f"{API_BASE_URL}/Split/Runs",
            headers=build_headers(api_key),
            params={
                "runIds": run_ids,
                "splitLength": 5,
                "useYards": False,
                "includeRawPeaksAndAverages": True,
            },
            timeout=API_TIMEOUT_SECONDS,
        )
    except requests.RequestException as exc:
        return None, f"Communication error while loading split run data: {exc}"

    if response.status_code == 200:
        payload = response.json()
        if isinstance(payload, list):
            return payload, None
        return None, "The split runs response did not return the expected list format."

    if response.status_code in (401, 403):
        return None, "The API key is no longer authorized to load split run data."

    return None, f"1080 API returned an unexpected response while loading split run data: {response.status_code}"


def merge_split_payloads(payloads: list[dict[str, Any]]) -> dict[str, Any]:
    if not payloads:
        return {"reports": []}

    merged = dict(payloads[0])
    merged_reports: list[dict[str, Any]] = []
    for payload in payloads:
        reports = payload.get("reports") or []
        if isinstance(reports, list):
            merged_reports.extend(reports)
    merged["reports"] = merged_reports
    return merged


def build_split_collection_from_reports(
    reports: list[dict[str, Any]],
    source_entity_id: str,
    source_entity_type: str,
) -> dict[str, Any]:
    return {
        "sourceEntityId": source_entity_id,
        "sourceEntityType": source_entity_type,
        "reports": reports,
    }


def decode_sampledata_base64(sample_b64: str) -> list[dict[str, float]]:
    if not sample_b64:
        return []

    normalized = "".join(str(sample_b64).strip().split()).replace("-", "+").replace("_", "/")
    padding = len(normalized) % 4
    if padding:
        normalized += "=" * (4 - padding)

    try:
        raw = base64.b64decode(normalized, validate=False)
    except Exception:
        return []

    if len(raw) % 20 != 0:
        return []

    rows: list[dict[str, float]] = []
    unpackers = ("<fffff", ">fffff")
    for fmt in unpackers:
        rows.clear()
        try:
            for time_s, position_m, speed_mps, acceleration_mps2, force_n in struct.iter_unpack(fmt, raw):
                rows.append(
                    {
                        "time_s": float(time_s),
                        "position_m": float(position_m),
                        "speed_mps": float(speed_mps),
                        "acceleration_mps2": float(acceleration_mps2),
                        "force_n": float(force_n),
                    }
                )
        except struct.error:
            continue

        if rows:
            return rows.copy()

    return []


def compute_deceleration_profile_from_samples(
    samples: list[dict[str, float]],
    *,
    v_stop: float = DECEL_V_STOP,
    acc_threshold: float = DECEL_ACC_THRESHOLD,
    min_points_after_start: int = 15,
    min_t_after_start: float = 0.4,
    end_speed_tol: float = 0.3,
    min_tts: float = 0.25,
    min_dts: float = 0.8,
    min_decmax: float = -1.0,
) -> dict[str, Any]:
    ordered_samples = sorted(samples, key=lambda sample: float(sample.get("time_s") or 0))
    if not ordered_samples:
        return {"ok": False, "reason": "empty"}

    peak_index = max(
        range(len(ordered_samples)),
        key=lambda index: float(ordered_samples[index].get("speed_mps") or 0),
    )
    start_index = next(
        (
            index
            for index, sample in enumerate(ordered_samples[peak_index:], start=peak_index)
            if float(sample.get("acceleration_mps2") or 0) <= acc_threshold
        ),
        None,
    )
    if start_index is None:
        return {"ok": False, "reason": "no_decel_threshold"}

    if (len(ordered_samples) - 1 - start_index) < min_points_after_start:
        return {"ok": False, "reason": "too_few_points_after_start"}

    start_sample = ordered_samples[start_index]
    end_sample = ordered_samples[-1]
    start_time = float(start_sample.get("time_s") or 0)
    if (float(end_sample.get("time_s") or 0) - start_time) < min_t_after_start:
        return {"ok": False, "reason": "too_short_after_start"}

    stop_index = next(
        (
            index
            for index in range(start_index, len(ordered_samples))
            if float(ordered_samples[index].get("speed_mps") or 0) <= v_stop
        ),
        None,
    )
    if stop_index is None:
        return {"ok": False, "reason": "no_stop_reached"}

    stop_sample = ordered_samples[stop_index]
    time_to_stop = float(stop_sample.get("time_s") or 0) - start_time
    distance_to_stop = float(stop_sample.get("position_m") or 0) - float(start_sample.get("position_m") or 0)
    if time_to_stop <= 0:
        return {"ok": False, "reason": "bad_tts"}
    if time_to_stop < min_tts:
        return {"ok": False, "reason": "tts_too_small"}
    if distance_to_stop < min_dts:
        return {"ok": False, "reason": "dts_too_small"}

    decel_segment = ordered_samples[start_index:stop_index + 1]
    if not decel_segment:
        return {"ok": False, "reason": "empty_segment"}

    deceleration_max = min(float(sample.get("acceleration_mps2") or 0) for sample in decel_segment)
    if deceleration_max > min_decmax:
        return {"ok": False, "reason": "not_enough_braking"}

    end_speed = float(end_sample.get("speed_mps") or 0)
    if end_speed > (v_stop + end_speed_tol):
        return {"ok": False, "reason": "end_speed_too_high"}

    negative_values = [
        float(sample.get("acceleration_mps2") or 0)
        for sample in decel_segment
        if float(sample.get("acceleration_mps2") or 0) < 0
    ]
    if not negative_values:
        return {"ok": False, "reason": "no_negative_acc_values"}

    average_deceleration = -sum(negative_values) / len(negative_values)
    deceleration_max_abs = -deceleration_max
    top_speed = float(ordered_samples[peak_index].get("speed_mps") or 0)
    segment_start_speed = float(start_sample.get("speed_mps") or 0)
    mid_velocity = 0.5 * (segment_start_speed + v_stop)
    mid_relative_index = next(
        (
            index
            for index, sample in enumerate(decel_segment)
            if float(sample.get("speed_mps") or 0) <= mid_velocity
        ),
        min(
            range(len(decel_segment)),
            key=lambda index: abs(float(decel_segment[index].get("speed_mps") or 0) - mid_velocity),
        ),
    )
    plot_segment = []
    for sample in ordered_samples:
        plot_segment.append(
            {
                "time_s": float(sample.get("time_s") or 0),
                "t_rel": float(sample.get("time_s") or 0) - start_time,
                "speed_mps": float(sample.get("speed_mps") or 0),
                "acceleration_mps2": float(sample.get("acceleration_mps2") or 0),
            }
        )

    return {
        "ok": True,
        "averageDeceleration": average_deceleration,
        "DecM": deceleration_max_abs,
        "VMax": top_speed,
        "TTS": time_to_stop,
        "DTS": distance_to_stop,
        "vStart": segment_start_speed,
        "startIndex": start_index,
        "stopIndex": stop_index,
        "midIndex": start_index + mid_relative_index,
        "plotSamples": plot_segment,
        "vStop": v_stop,
        "exerciseName": "",
    }


def build_vmax_stop_plot_samples_from_raw(
    samples: list[dict[str, float]],
    *,
    v_stop: float = DECEL_V_STOP,
) -> list[dict[str, float]]:
    ordered_samples = sorted(samples, key=lambda sample: float(sample.get("time_s") or 0))
    if not ordered_samples:
        return []

    peak_index = max(
        range(len(ordered_samples)),
        key=lambda index: float(ordered_samples[index].get("speed_mps") or 0),
    )
    stop_index = next(
        (
            index
            for index in range(peak_index, len(ordered_samples))
            if float(ordered_samples[index].get("speed_mps") or 0) <= v_stop
        ),
        len(ordered_samples) - 1,
    )
    if stop_index <= peak_index:
        return []

    segment = ordered_samples[peak_index:stop_index + 1]
    start_time = float(segment[0].get("time_s") or 0)
    return [
        {
            "time_s": float(sample.get("time_s") or 0),
            "t_rel": float(sample.get("time_s") or 0) - start_time,
            "speed_mps": float(sample.get("speed_mps") or 0),
            "acceleration_mps2": float(sample.get("acceleration_mps2") or 0),
        }
        for sample in segment
    ]


def build_derived_split_runs_from_training_data(set_payloads: list[dict[str, Any]]) -> list[dict[str, Any]]:
    derived_runs: list[dict[str, Any]] = []

    for set_payload in set_payloads:
        motion_groups = set_payload.get("motionGroups") or []
        for motion_group in motion_groups:
            motions = motion_group.get("motions") or []
            if not motions:
                continue

            total_distance = sum(float(motion.get("totalDistance") or 0) for motion in motions)
            total_time = sum(float(motion.get("totalTime") or 0) for motion in motions)
            top_speed = max((float(motion.get("topSpeed") or 0) for motion in motions), default=0.0)
            max_acceleration = max(
                (
                    float((motion.get("accelDecelStats") or {}).get("accelerationMax") or 0)
                    for motion in motions
                ),
                default=0.0,
            )
            max_deceleration = max(
                (
                    float((motion.get("accelDecelStats") or {}).get("decelerationMax") or 0)
                    for motion in motions
                ),
                default=0.0,
            )
            deceleration_time = sum(
                float((motion.get("accelDecelStats") or {}).get("decelerationTime") or 0)
                for motion in motions
            )
            first_resistance = (motions[0].get("resistanceValues") or {}) if motions else {}
            load = first_resistance.get("concentricLoad")
            motion_details = []
            cumulative_time = 0.0
            for motion in motions:
                motion_time = float(motion.get("totalTime") or 0)
                avg_speed = float(((motion.get("averageValues") or {}).get("speed")) or 0)
                top_motion_speed = float(motion.get("topSpeed") or 0)
                phase_name = str(motion.get("phaseName") or "")
                motion_details.append(
                    {
                        "phaseName": phase_name,
                        "duration": motion_time,
                        "avgSpeed": avg_speed,
                        "topSpeed": top_motion_speed,
                        "startTime": cumulative_time,
                        "endTime": cumulative_time + motion_time,
                    }
                )
                cumulative_time += motion_time

            derived_runs.append(
                {
                    "motionGroupId": str(motion_group.get("id") or ""),
                    "distance": total_distance,
                    "time": total_time,
                    "topSpeed": top_speed,
                    "load": load,
                    "maxAcceleration": max_acceleration,
                    "maxDeceleration": max_deceleration,
                    "decelerationTime": deceleration_time,
                    "motions": motion_details,
                }
            )

    return derived_runs


def build_deceleration_runs_from_training_data(
    set_payloads: list[dict[str, Any]]
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    deceleration_runs: list[dict[str, Any]] = []
    raw_failures: list[dict[str, Any]] = []

    for set_payload in set_payloads:
        exercise_name = str(set_payload.get("exerciseName") or "")
        for motion_group in set_payload.get("motionGroups") or []:
            motions = motion_group.get("motions") or []
            if not motions:
                continue

            selected_motion = next(
                (
                    motion
                    for motion in motions
                    if motion.get("sampleData")
                ),
                None,
            )
            if not selected_motion:
                continue

            sample_data_raw = str(selected_motion.get("sampleData") or "")
            samples = decode_sampledata_base64(sample_data_raw)
            metrics = compute_deceleration_profile_from_samples(samples)
            if not metrics.get("ok"):
                raw_failures.append(
                    {
                        "motionGroupId": str(motion_group.get("id") or ""),
                        "exerciseName": exercise_name,
                        "created": selected_motion.get("created") or motion_group.get("created") or set_payload.get("created"),
                        "hasSampleData": bool(sample_data_raw),
                        "sampleDataLength": len(sample_data_raw),
                        "decodedSampleCount": len(samples),
                        "reason": metrics.get("reason") or "unknown",
                        "peakSpeed": max(
                            (float(sample.get("speed_mps") or 0) for sample in samples),
                            default=0.0,
                        ),
                        "endSpeed": float((samples[-1].get("speed_mps") or 0)) if samples else None,
                        "minAcceleration": min(
                            (float(sample.get("acceleration_mps2") or 0) for sample in samples),
                            default=None,
                        ),
                        "rawPlotSampleCount": len(build_vmax_stop_plot_samples_from_raw(samples)),
                    }
                )
                continue

            metrics["motionGroupId"] = str(motion_group.get("id") or "")
            metrics["exerciseName"] = exercise_name
            metrics["created"] = selected_motion.get("created") or motion_group.get("created") or set_payload.get("created")
            metrics["_debug"] = {
                "source": "raw_samples",
                "hasSampleData": bool(sample_data_raw),
                "sampleDataLength": len(sample_data_raw),
                "decodedSampleCount": len(samples),
            }
            deceleration_runs.append(metrics)

    return deceleration_runs, raw_failures


def build_fallback_deceleration_runs(
    split_reports: list[dict[str, Any]],
    set_payloads: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    if not split_reports or not set_payloads:
        return []

    motion_group_lookup: dict[str, dict[str, Any]] = {}
    exercise_name_lookup: dict[str, str] = {}
    for set_payload in set_payloads:
        exercise_name = str(set_payload.get("exerciseName") or "")
        for motion_group in set_payload.get("motionGroups") or []:
            motion_group_id = str(motion_group.get("id") or "")
            if not motion_group_id:
                continue
            motion_group_lookup[motion_group_id] = motion_group
            exercise_name_lookup[motion_group_id] = exercise_name

    fallback_runs: list[dict[str, Any]] = []
    for report in split_reports:
        motion_group_id = str(report.get("motionGroupId") or "")
        motion_group = motion_group_lookup.get(motion_group_id)
        if not motion_group:
            continue

        motions = motion_group.get("motions") or []
        if not motions:
            continue
        motion = motions[0]
        accel_decel_stats = motion.get("accelDecelStats") or {}
        top_speed = float(motion.get("topSpeed") or 0)
        deceleration_time = float(accel_decel_stats.get("decelerationTime") or 0)
        max_deceleration = float(accel_decel_stats.get("decelerationMax") or 0)
        total_distance = float(motion.get("totalDistance") or 0)
        top_speed_position = float(accel_decel_stats.get("topSpeedPosition") or 0)
        distance_to_stop = max(total_distance - top_speed_position, 0.0)

        if deceleration_time <= 0 or max_deceleration <= 0 or top_speed <= 0:
            continue

        average_deceleration = max((top_speed - DECEL_V_STOP) / deceleration_time, 0.0)
        sample_data_raw = str(motion.get("sampleData") or "")
        raw_samples = decode_sampledata_base64(sample_data_raw) if sample_data_raw else []
        plot_samples = build_vmax_stop_plot_samples_from_raw(raw_samples)
        plot_source = "fallback_from_raw_samples"

        if not plot_samples:
            plot_source = "fallback_from_split"
            anchor_samples: list[dict[str, float]] = []
            elapsed_time = 0.0
            splits = report.get("splits") or []
            first_speed = float((splits[0] or {}).get("topSpeed") or top_speed) if splits else top_speed
            anchor_samples.append(
                {
                    "time_s": 0.0,
                    "t_rel": 0.0,
                    "speed_mps": first_speed,
                    "acceleration_mps2": 0.0,
                }
            )
            for split in splits:
                split_time = float(split.get("time") or 0)
                top_split_speed = float(split.get("topSpeed") or 0)
                elapsed_time += split_time
                anchor_samples.append(
                    {
                        "time_s": elapsed_time,
                        "t_rel": elapsed_time,
                        "speed_mps": top_split_speed,
                        "acceleration_mps2": 0.0,
                    }
                )
            final_time = max(elapsed_time, deceleration_time)
            if not anchor_samples or anchor_samples[-1]["speed_mps"] > DECEL_V_STOP:
                anchor_samples.append(
                    {
                        "time_s": final_time,
                        "t_rel": final_time,
                        "speed_mps": DECEL_V_STOP,
                        "acceleration_mps2": 0.0,
                    }
                )

            plot_samples = []
            for index, current in enumerate(anchor_samples):
                plot_samples.append(dict(current))
                if index == len(anchor_samples) - 1:
                    continue
                next_sample = anchor_samples[index + 1]
                subdivisions = 5
                for step in range(1, subdivisions):
                    ratio = step / subdivisions
                    interpolated_time = current["t_rel"] + (next_sample["t_rel"] - current["t_rel"]) * ratio
                    interpolated_speed = current["speed_mps"] + (next_sample["speed_mps"] - current["speed_mps"]) * ratio
                    plot_samples.append(
                        {
                            "time_s": interpolated_time,
                            "t_rel": interpolated_time,
                            "speed_mps": interpolated_speed,
                            "acceleration_mps2": 0.0,
                        }
                    )

        if plot_samples:
            decm_time = min(deceleration_time, plot_samples[-1]["t_rel"])
            nearest_index = min(
                range(len(plot_samples)),
                key=lambda index: abs(plot_samples[index]["t_rel"] - decm_time),
            )
            plot_samples[nearest_index]["acceleration_mps2"] = -max_deceleration

        mid_index = len(plot_samples) // 2 if plot_samples else 0
        fallback_runs.append(
            {
                "motionGroupId": motion_group_id,
                "exerciseName": exercise_name_lookup.get(motion_group_id, ""),
                "created": motion.get("created") or motion_group.get("created"),
                "averageDeceleration": average_deceleration,
                "DecM": max_deceleration,
                "VMax": top_speed,
                "TTS": deceleration_time,
                "DTS": distance_to_stop,
                "vStart": top_speed,
                "vStop": DECEL_V_STOP,
                "startIndex": 0,
                "stopIndex": max(len(plot_samples) - 1, 0),
                "midIndex": mid_index,
                "plotSamples": plot_samples,
                "isFallback": True,
                "_debug": {
                    "source": plot_source,
                    "hasSampleData": bool(sample_data_raw),
                    "sampleDataLength": len(sample_data_raw),
                    "decodedSampleCount": len(raw_samples),
                    "splitPointCount": len(plot_samples),
                },
            }
        )

    return fallback_runs


def append_split_debug(
    debug_steps: list[dict[str, Any]],
    step: str,
    *,
    source_id: str = "",
    payload: Any = None,
    error: str | None = None,
) -> None:
    debug_steps.append(
        {
            "step": step,
            "sourceId": source_id,
            "error": error,
            "payload": payload,
        }
    )


def sanitize_for_debug(value: Any) -> Any:
    if isinstance(value, dict):
        return {
            key: sanitize_for_debug(item)
            for key, item in value.items()
            if key != "_debug_fetches"
        }
    if isinstance(value, list):
        return [sanitize_for_debug(item) for item in value]
    return value


def load_split_payload_for_exercise(
    api_key: str,
    exercise_id: str,
    exercise: dict[str, Any] | None = None,
) -> tuple[dict[str, Any] | None, str | None]:
    debug_steps: list[dict[str, Any]] = []
    training_set_payloads: list[dict[str, Any]] = []
    exercise_payload, error = fetch_split_exercise(api_key, exercise_id)
    append_split_debug(
        debug_steps,
        "Split/Exercise",
        source_id=exercise_id,
        payload=exercise_payload,
        error=error,
    )
    if error:
        return None, error

    exercise_has_reports = bool(exercise_payload.get("reports") or [])

    set_payloads: list[dict[str, Any]] = []
    set_errors: list[str] = []
    for exercise_set in (exercise or {}).get("sets") or []:
        set_id = str(exercise_set.get("id") or "")
        if not set_id:
            continue
        set_training_payload, set_training_error = fetch_training_data_set(api_key, set_id)
        append_split_debug(
            debug_steps,
            "TrainingData/Set",
            source_id=set_id,
            payload=set_training_payload,
            error=set_training_error,
        )
        if set_training_error:
            set_errors.append(set_training_error)
            continue
        if set_training_payload:
            training_set_payloads.append(set_training_payload)

        if exercise_has_reports:
            continue

        motion_groups = (set_training_payload or {}).get("motionGroups") or []
        run_ids = [
            str(group.get("id") or "")
            for group in motion_groups
            if str(group.get("id") or "")
        ]
        if run_ids:
            run_reports, run_error = fetch_split_runs(api_key, run_ids)
            append_split_debug(
                debug_steps,
                "Split/Runs",
                source_id=set_id,
                payload={"runIds": run_ids, "reports": run_reports},
                error=run_error,
            )
            if run_error:
                set_errors.append(run_error)
                continue
            if run_reports:
                set_payloads.append(
                    build_split_collection_from_reports(run_reports, set_id, "Set")
                )
                continue

        set_payload, set_error = fetch_split_set(api_key, set_id)
        append_split_debug(
            debug_steps,
            "Split/Set",
            source_id=set_id,
            payload=set_payload,
            error=set_error,
        )
        if set_error:
            set_errors.append(set_error)
            continue
        if set_payload:
            set_payloads.append(set_payload)

    response_payload = exercise_payload if exercise_has_reports else merge_split_payloads(set_payloads)
    derived_runs = build_derived_split_runs_from_training_data(training_set_payloads)
    deceleration_runs, raw_failures = build_deceleration_runs_from_training_data(training_set_payloads)
    if not deceleration_runs:
        deceleration_runs = build_fallback_deceleration_runs(
            response_payload.get("reports") or [],
            training_set_payloads,
        )
    response_payload["_debug_fetches"] = debug_steps
    response_payload["_derived_runs"] = derived_runs
    response_payload["_deceleration_runs"] = deceleration_runs
    response_payload["_deceleration_raw_failures"] = raw_failures
    if response_payload.get("reports"):
        return response_payload, None

    if set_errors:
        return response_payload, set_errors[0]

    return response_payload, None


def storage_namespace(api_key: str) -> str:
    digest = hashlib.sha256(api_key.encode("utf-8")).hexdigest()[:12]
    return f"1080_clients:{digest}"


def recent_opened_sessions_namespace(api_key: str) -> str:
    digest = hashlib.sha256(api_key.encode("utf-8")).hexdigest()[:12]
    return f"1080_recent_opened_sessions:{digest}"


def normalize_text(text: str) -> str:
    return unicodedata.normalize("NFD", str(text)).encode("ascii", "ignore").decode("utf-8").lower().strip()


def safe_filename(value: str) -> str:
    value = normalize_text(value).replace(" ", "_")
    value = "".join(ch for ch in value if ch.isalnum() or ch in {"_", "-"})
    return value[:120] if value else "export"


def ensure_uploaded_logos_dir() -> None:
    UPLOADED_LOGOS_DIR.mkdir(exist_ok=True)


def list_uploaded_logos() -> list[str]:
    ensure_uploaded_logos_dir()
    valid_suffixes = {".png", ".jpg", ".jpeg", ".webp"}
    return sorted(
        [
            file.name
            for file in UPLOADED_LOGOS_DIR.iterdir()
            if file.is_file() and file.suffix.lower() in valid_suffixes
        ]
    )


def save_uploaded_logo(uploaded_file: Any) -> str:
    ensure_uploaded_logos_dir()
    original_name = getattr(uploaded_file, "name", "logo.png")
    stem = Path(original_name).stem
    suffix = Path(original_name).suffix.lower() or ".png"
    base_name = safe_filename(stem) or "logo"
    candidate = f"{base_name}{suffix}"
    target_path = UPLOADED_LOGOS_DIR / candidate

    if target_path.exists():
        raise FileExistsError(candidate)

    target_path.write_bytes(uploaded_file.getvalue())
    return candidate


def load_saved_logo_bytes(file_name: str) -> bytes | None:
    ensure_uploaded_logos_dir()
    if not file_name:
        return None

    path = UPLOADED_LOGOS_DIR / file_name
    if not path.is_file():
        return None

    return path.read_bytes()


def render_saved_logo_picker(
    saved_logos: list[str],
    selected_logo_state_key: str,
    language: str = "English",
) -> str:
    texts = PDF_TEXT[language]
    choose_logo_option = "__choose_logo__"
    no_saved_logo_option = "__no_saved_logo__"
    current_selection = st.session_state.get(selected_logo_state_key, choose_logo_option)
    current_logo_bytes = (
        load_saved_logo_bytes(current_selection)
        if current_selection not in {no_saved_logo_option, choose_logo_option}
        else None
    )

    if current_logo_bytes:
        preview_col, label_col = st.columns([0.2, 0.8], vertical_alignment="center")
        preview_col.image(current_logo_bytes, width=32)
        label_col.caption(f"{texts['selected_logo']}: {current_selection}")
    else:
        st.caption(texts["no_logo_selected"])

    with st.popover(texts["choose_logo"]):
        for file_name in saved_logos:
            row_col1, row_col2 = st.columns([0.22, 0.78], vertical_alignment="center")
            logo_bytes = load_saved_logo_bytes(file_name)
            if logo_bytes:
                row_col1.image(logo_bytes, width=28)
            else:
                row_col1.caption("N/A")

            button_label = file_name
            if file_name == current_selection:
                button_label = f"{file_name}  [selected]"

            if row_col2.button(
                button_label,
                key=f"{selected_logo_state_key}_{file_name}",
                width="stretch",
                type="primary" if file_name == current_selection else "secondary",
            ):
                st.session_state[selected_logo_state_key] = file_name

    return st.session_state.get(selected_logo_state_key, choose_logo_option)


def ensure_uploaded_player_photos_dir() -> None:
    UPLOADED_PLAYER_PHOTOS_DIR.mkdir(exist_ok=True)


def save_player_photo(uploaded_file: Any, client_id: str) -> str:
    ensure_uploaded_player_photos_dir()
    suffix = Path(getattr(uploaded_file, "name", "photo.png")).suffix.lower() or ".png"
    base_name = safe_filename(client_id) or "client"

    for existing in UPLOADED_PLAYER_PHOTOS_DIR.glob(f"{base_name}.*"):
        if existing.is_file():
            existing.unlink()

    file_name = f"{base_name}{suffix}"
    target_path = UPLOADED_PLAYER_PHOTOS_DIR / file_name
    target_path.write_bytes(uploaded_file.getvalue())
    return file_name


def get_saved_player_photo_name(client_id: str) -> str | None:
    ensure_uploaded_player_photos_dir()
    base_name = safe_filename(client_id) or "client"
    for file in sorted(UPLOADED_PLAYER_PHOTOS_DIR.glob(f"{base_name}.*")):
        if file.is_file():
            return file.name
    return None


def load_player_photo_bytes(client_id: str) -> bytes | None:
    saved_name = get_saved_player_photo_name(client_id)
    if not saved_name:
        return None
    path = UPLOADED_PLAYER_PHOTOS_DIR / saved_name
    if not path.is_file():
        return None
    return path.read_bytes()


def validate_uploaded_file_size(uploaded_file: Any, max_bytes: int, label: str) -> bool:
    if uploaded_file is None:
        return True

    file_size = len(uploaded_file.getvalue())
    if file_size > max_bytes:
        max_mb = max_bytes / (1024 * 1024)
        st.error(f"{label} is too large. Maximum allowed size is {max_mb:.0f} MB.")
        return False

    return True


def render_logo_library_selector(key_prefix: str, language: str = "English") -> bytes | None:
    texts = PDF_TEXT[language]
    saved_logos = list_uploaded_logos()
    selected_logo_state_key = f"{key_prefix}_selected_logo_name"
    pending_logo_state_key = f"{key_prefix}_pending_saved_logo"
    processed_logo_upload_key = f"{key_prefix}_processed_logo_upload"
    no_saved_logo_option = "__no_saved_logo__"
    choose_logo_option = "__choose_logo__"
    default_option = no_saved_logo_option if not saved_logos else choose_logo_option

    if selected_logo_state_key not in st.session_state:
        st.session_state[selected_logo_state_key] = default_option

    pending_logo_name = st.session_state.pop(pending_logo_state_key, None)
    if pending_logo_name:
        st.session_state[selected_logo_state_key] = pending_logo_name

    options = [choose_logo_option, *saved_logos] if saved_logos else [no_saved_logo_option]
    if st.session_state[selected_logo_state_key] not in options:
        st.session_state[selected_logo_state_key] = default_option

    st.markdown(f"##### {texts['branding']}")
    select_col, upload_col = st.columns([1, 1])
    with select_col:
        if saved_logos:
            selected_logo = render_saved_logo_picker(saved_logos, selected_logo_state_key, language)
            if selected_logo not in {no_saved_logo_option, choose_logo_option}:
                if st.button(
                    texts["clear_selection"],
                    key=f"{selected_logo_state_key}_clear_inline",
                    width="stretch",
                ):
                    st.session_state[selected_logo_state_key] = choose_logo_option
                    st.rerun()
        else:
            st.selectbox(
                texts["choose_logo"],
                options=options,
                format_func=lambda option: texts["no_logo_selected"] if option == no_saved_logo_option else texts["choose_logo"],
                key=selected_logo_state_key,
                disabled=True,
            )
            selected_logo = no_saved_logo_option
    uploaded_logo = upload_col.file_uploader(
        texts["upload_new_logo"],
        type=["png", "jpg", "jpeg", "webp"],
        accept_multiple_files=False,
        key=f"{key_prefix}_logo_upload",
        help="Maximum file size: 2 MB",
    )

    logo_is_valid = validate_uploaded_file_size(uploaded_logo, MAX_LOGO_BYTES, "Logo")

    if uploaded_logo is None:
        st.session_state.pop(processed_logo_upload_key, None)
    elif logo_is_valid:
        upload_token = f"{getattr(uploaded_logo, 'name', 'logo')}:{len(uploaded_logo.getvalue())}"
        if st.session_state.get(processed_logo_upload_key) != upload_token:
            st.session_state[processed_logo_upload_key] = upload_token
            try:
                saved_name = save_uploaded_logo(uploaded_logo)
            except FileExistsError as exc:
                st.error(f"Logo `{exc.args[0]}` has already been uploaded.")
            else:
                st.session_state[pending_logo_state_key] = saved_name
                st.success(f"Saved logo: {saved_name}")
                st.rerun()

    if uploaded_logo is not None and logo_is_valid:
        return uploaded_logo.getvalue()

    if selected_logo not in {no_saved_logo_option, choose_logo_option}:
        return load_saved_logo_bytes(selected_logo)

    return None


def render_player_photo_selector(key_prefix: str, client_id: str, language: str = "English") -> bytes | None:
    texts = PDF_TEXT[language]
    saved_photo = get_saved_player_photo_name(client_id)
    current_bytes = load_player_photo_bytes(client_id)

    if current_bytes:
        st.caption(texts["player_photo"])
        thumb_col, _ = st.columns([0.34, 0.66])
        thumb_col.image(current_bytes, width=140)
        return current_bytes

    st.caption(texts["add_player_photo"])
    uploaded_photo = st.file_uploader(
        texts["upload_player_photo"],
        type=["png", "jpg", "jpeg", "webp"],
        accept_multiple_files=False,
        key=f"{key_prefix}_player_photo_upload",
        help="Maximum file size: 5 MB",
    )

    photo_is_valid = validate_uploaded_file_size(uploaded_photo, MAX_PLAYER_PHOTO_BYTES, "Player photo")
    if uploaded_photo is not None and photo_is_valid:
        saved_name = save_player_photo(uploaded_photo, client_id)
        st.success(f"Saved player photo: {saved_name}")
        st.rerun()

    return current_bytes


def iso_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def prune_recent_opened_sessions(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    cutoff = datetime.now(timezone.utc) - RECENT_OPENED_SESSION_DURATION
    pruned: list[dict[str, Any]] = []

    for item in items:
        try:
            opened_at_raw = str(item.get("openedAt") or "")
            opened_at = datetime.fromisoformat(opened_at_raw)
        except ValueError:
            continue

        if opened_at.tzinfo is None:
            opened_at = opened_at.replace(tzinfo=timezone.utc)

        if opened_at >= cutoff:
            pruned.append(item)

    pruned.sort(key=lambda item: str(item.get("openedAt") or ""), reverse=True)
    return pruned[:10]


def format_sync_label(value: str) -> str:
    if not value:
        return "Never"

    try:
        timestamp = datetime.fromisoformat(value)
    except ValueError:
        return value

    return timestamp.astimezone().strftime("%Y-%m-%d %H:%M:%S %Z")


def handle_auth_storage_bridge() -> None:
    storage_result = CLIENT_STORAGE_COMPONENT(
        data={
            "storageKey": "1080_auth_bridge_clients",
            "command": "read",
            "commandId": "",
            "clientsJson": "",
            "lastSynced": "",
            "authStorageKey": AUTH_STORAGE_KEY,
            "authCommand": st.session_state["auth_storage_command"],
            "authCommandId": st.session_state["auth_storage_command_id"],
            "authSessionToken": st.session_state["auth_session_token"],
        },
        default={"clients_json": "", "last_synced": "", "auth_session_token": "", "auth_storage_ready": False},
        on_clients_json_change=lambda: None,
        on_last_synced_change=lambda: None,
        on_auth_session_token_change=lambda: None,
        on_auth_storage_ready_change=lambda: None,
        key="auth_storage_bridge",
        height=0,
    )

    auth_token = getattr(storage_result, "auth_session_token", "") or ""
    auth_storage_ready = bool(getattr(storage_result, "auth_storage_ready", False))

    if (
        not st.session_state["auth_verified"]
        and not st.session_state["auth_storage_autoload_complete"]
    ):
        if not auth_storage_ready:
            return

        if auth_token and restore_auth_session_from_token(auth_token):
            st.rerun()

        st.session_state["auth_storage_autoload_complete"] = True
        if auth_token:
            queue_auth_storage_command("clear")


def ensure_session_defaults() -> None:
    st.session_state.setdefault("auth_verified", False)
    st.session_state.setdefault("auth_user_email", "")
    st.session_state.setdefault("api_key", "")
    st.session_state.setdefault("api_valid", False)
    st.session_state.setdefault("auth_storage_command", "read")
    st.session_state.setdefault("auth_storage_command_id", "")
    st.session_state.setdefault("auth_session_token", "")
    st.session_state.setdefault("auth_storage_autoload_complete", False)
    st.session_state.setdefault("clients_cache", [])
    st.session_state.setdefault("clients_last_synced", "")
    st.session_state.setdefault("client_storage_command", "read")
    st.session_state.setdefault("client_storage_command_id", "")
    st.session_state.setdefault("client_storage_payload", "")
    st.session_state.setdefault("client_storage_autoload_complete", False)
    st.session_state.setdefault("client_storage_error", "")
    st.session_state.setdefault("recent_opened_sessions", [])
    st.session_state.setdefault("recent_opened_command", "read")
    st.session_state.setdefault("recent_opened_command_id", "")
    st.session_state.setdefault("recent_opened_sessions_payload", "")
    st.session_state.setdefault("recent_opened_autoload_complete", False)
    st.session_state.setdefault("selected_client_id", "")
    st.session_state.setdefault("selected_client_last_id", "")
    st.session_state.setdefault("client_sessions", [])
    st.session_state.setdefault("client_sessions_error", "")
    st.session_state.setdefault("client_sessions_loaded_for", "")
    st.session_state.setdefault("session_filter_from", date.today() - timedelta(days=7))
    st.session_state.setdefault("session_filter_to", date.today())
    st.session_state.setdefault("session_filter_last_from", date.today() - timedelta(days=7))
    st.session_state.setdefault("session_filter_last_to", date.today())
    st.session_state.setdefault("selected_session_id", "")
    st.session_state.setdefault("session_detail", None)
    st.session_state.setdefault("session_detail_error", "")
    st.session_state.setdefault("exercise_report_cache", {})
    st.session_state.setdefault("exercise_report_errors", {})


def sync_clients_to_storage(clients: list[dict[str, Any]], last_synced: str) -> None:
    st.session_state["clients_cache"] = clients
    st.session_state["clients_last_synced"] = last_synced
    st.session_state["client_storage_command"] = "write"
    st.session_state["client_storage_command_id"] = f"write:{last_synced}"
    st.session_state["client_storage_payload"] = json.dumps(clients)
    st.session_state["client_storage_error"] = ""


def sync_recent_opened_sessions_to_storage(items: list[dict[str, Any]]) -> None:
    pruned_items = prune_recent_opened_sessions(items)
    st.session_state["recent_opened_sessions"] = pruned_items
    st.session_state["recent_opened_command"] = "write"
    st.session_state["recent_opened_command_id"] = f"write:{iso_now()}"
    st.session_state["recent_opened_sessions_payload"] = json.dumps(pruned_items)
    st.session_state["recent_opened_autoload_complete"] = True


def remember_opened_session(session_detail: dict[str, Any]) -> None:
    session_id = str(session_detail.get("id") or "")
    if not session_id:
        return

    current_items = st.session_state.get("recent_opened_sessions", [])
    existing = [
        item for item in current_items
        if str(item.get("id") or "") != session_id
    ]
    existing.insert(
        0,
        {
            "id": session_id,
            "timestamp": session_detail.get("timestamp"),
            "clientId": session_detail.get("clientId"),
            "openedAt": iso_now(),
        },
    )
    sync_recent_opened_sessions_to_storage(existing)


def reset_client_cache_state() -> None:
    st.session_state["clients_cache"] = []
    st.session_state["clients_last_synced"] = ""
    st.session_state["client_storage_command"] = "clear"
    st.session_state["client_storage_command_id"] = f"clear:{iso_now()}"
    st.session_state["client_storage_payload"] = ""
    st.session_state["client_storage_autoload_complete"] = False
    st.session_state["client_storage_error"] = ""
    st.session_state["selected_client_id"] = ""
    st.session_state["selected_client_last_id"] = ""
    st.session_state["client_sessions"] = []
    st.session_state["client_sessions_error"] = ""
    st.session_state["client_sessions_loaded_for"] = ""
    st.session_state["session_filter_from"] = date.today() - timedelta(days=7)
    st.session_state["session_filter_to"] = date.today()
    st.session_state["session_filter_last_from"] = date.today() - timedelta(days=7)
    st.session_state["session_filter_last_to"] = date.today()
    st.session_state["selected_session_id"] = ""
    st.session_state["session_detail"] = None
    st.session_state["session_detail_error"] = ""
    st.session_state["exercise_report_cache"] = {}
    st.session_state["exercise_report_errors"] = {}


def logout() -> None:
    st.session_state["auth_verified"] = False
    st.session_state["auth_user_email"] = ""
    st.session_state["api_key"] = ""
    st.session_state["api_valid"] = False
    st.session_state["auth_storage_autoload_complete"] = True
    queue_auth_storage_command("clear")
    reset_client_cache_state()


def get_auth_users() -> dict[str, str]:
    users = st.secrets.get("auth_users", {})
    return dict(users) if isinstance(users, Mapping) else {}


def verify_password(password: str, stored_hash: str) -> bool:
    try:
        algorithm, iterations_str, salt, expected_hex = stored_hash.split("$", 3)
        if algorithm != "pbkdf2_sha256":
            return False
        iterations = int(iterations_str)
    except (ValueError, TypeError):
        return False

    derived = hashlib.pbkdf2_hmac(
        "sha256",
        password.encode("utf-8"),
        salt.encode("utf-8"),
        iterations,
    ).hex()
    return hmac.compare_digest(derived, expected_hex)


def render_login() -> None:
    _, center_col, _ = st.columns([1.2, 1, 1.2])
    with center_col:
        st.markdown(
            """
            <div class="hero-card">
              <h1 style="margin:0;">1080 Reports</h1>
              <p class="section-intro">Sign in with your email and password to access reporting tools.</p>
            </div>
            """,
            unsafe_allow_html=True,
        )

        auth_users = get_auth_users()
        if not auth_users:
            st.error("No app users are configured. Add `auth_users` to Streamlit secrets.")
            return
        if not APP_API_KEY:
            st.error("1080 API key is not configured. Set `api_1080_key` in Streamlit secrets or `API1080_KEY` as an environment variable.")
            return

        with st.form("auth-login-form"):
            email = st.text_input("Email", placeholder="you@example.com").strip().lower()
            password = st.text_input("Password", type="password")
            submitted = st.form_submit_button("Sign in", width="stretch")

        if submitted:
            stored_hash = auth_users.get(email)
            if not stored_hash or not verify_password(password, str(stored_hash)):
                st.error("Incorrect email or password.")
                return

            with st.spinner("Validating 1080 API access..."):
                is_valid, message = validate_api_key(APP_API_KEY)

            if is_valid:
                st.session_state["auth_verified"] = True
                st.session_state["auth_user_email"] = email
                st.session_state["api_key"] = APP_API_KEY
                st.session_state["api_valid"] = True
                if AUTH_SESSION_SECRET:
                    queue_auth_storage_command("write", create_auth_session_token(email))
                st.session_state["client_storage_autoload_complete"] = False
                st.session_state["auth_storage_autoload_complete"] = True
                st.rerun()

            st.session_state["api_valid"] = False
            st.error(message)


def load_clients_from_api(api_key: str) -> bool:
    clients, error = fetch_clients(api_key)
    if error:
        st.session_state["client_storage_error"] = error
        return False

    sync_clients_to_storage(clients or [], iso_now())
    return True


def format_optional_value(value: Any) -> str:
    if value in (None, "", []):
        return "-"
    if isinstance(value, list):
        return ", ".join(str(item) for item in value) if value else "-"
    return str(value)


def format_session_timestamp(value: Any) -> str:
    if not value:
        return "-"

    try:
        normalized = str(value).replace("Z", "+00:00")
        timestamp = datetime.fromisoformat(normalized)
    except ValueError:
        return str(value)

    return timestamp.astimezone().strftime("%Y-%m-%d %H:%M:%S %Z")


def get_client_lookup(clients: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    return {str(client.get("id") or ""): client for client in clients}


def get_client_display_name(client_id: Any, client_lookup: dict[str, dict[str, Any]]) -> str:
    client = client_lookup.get(str(client_id or ""))
    if not client:
        return "-"
    return format_optional_value(client.get("displayName"))


def has_missing_client_names(sessions: list[dict[str, Any]], client_lookup: dict[str, dict[str, Any]]) -> bool:
    for session in sessions:
        client_id = str(session.get("clientId") or "")
        if client_id and client_id not in client_lookup:
            return True
    return False


def ensure_session_scope_defaults(scope: str) -> None:
    st.session_state.setdefault(f"selected_session_id_{scope}", "")
    st.session_state.setdefault(f"session_detail_{scope}", None)
    st.session_state.setdefault(f"session_detail_error_{scope}", "")


def clear_session_scope(scope: str) -> None:
    st.session_state[f"selected_session_id_{scope}"] = ""
    st.session_state[f"session_detail_{scope}"] = None
    st.session_state[f"session_detail_error_{scope}"] = ""
    st.session_state["exercise_report_cache"] = {}
    st.session_state["exercise_report_errors"] = {}


def load_sessions_for_selected_client() -> bool:
    selected_client_id = st.session_state.get("selected_client_id", "")
    api_key = st.session_state.get("api_key", "")
    from_date = st.session_state.get("session_filter_from")
    to_date = st.session_state.get("session_filter_to")

    if not selected_client_id or not api_key:
        return False

    sessions, error = fetch_client_sessions(api_key, selected_client_id, from_date, to_date)
    if error:
        st.session_state["client_sessions"] = []
        st.session_state["client_sessions_error"] = error
        st.session_state["client_sessions_loaded_for"] = selected_client_id
        return False

    st.session_state["client_sessions"] = sessions or []
    st.session_state["client_sessions_error"] = ""
    st.session_state["client_sessions_loaded_for"] = selected_client_id
    clear_session_scope("client")
    return True


def load_selected_session_detail(scope: str) -> bool:
    api_key = st.session_state.get("api_key", "")
    session_id = st.session_state.get(f"selected_session_id_{scope}", "")

    if not api_key or not session_id:
        return False

    session_detail, error = fetch_session_detail(api_key, session_id)
    if error:
        st.session_state[f"session_detail_{scope}"] = None
        st.session_state[f"session_detail_error_{scope}"] = error
        return False

    st.session_state[f"session_detail_{scope}"] = session_detail
    st.session_state[f"session_detail_error_{scope}"] = ""
    st.session_state["exercise_report_cache"] = {}
    st.session_state["exercise_report_errors"] = {}
    if scope == "client":
        remember_opened_session(session_detail)
    return True


def format_decimal(value: Any, digits: int = 2) -> str:
    try:
        return f"{float(value):.{digits}f}"
    except (TypeError, ValueError):
        return "-"


def rounded_corner_cell(pdf: FPDF, x: float, y: float, w: float, h: float, text: str) -> None:
    pdf.set_xy(x, y)
    pdf.rect(x, y, w, h, round_corners=True, style="F")
    pdf.cell(w, h, text, 0, 0, "C")


def get_quadrant_badge_fill(quadrant: str) -> tuple[int, int, int]:
    if quadrant == "Q1":
        return (14, 108, 79)
    if quadrant == "Q2":
        return ORANGE_RGB
    if quadrant == "Q3":
        return RED_RGB
    return ORANGE_RGB


def get_quadrant_result_fill(quadrant: str) -> tuple[int, int, int]:
    if quadrant == "Q1":
        return GREEN_RGB
    if quadrant == "Q2":
        return (255, 188, 89)
    if quadrant == "Q3":
        return (240, 131, 133)
    return (255, 188, 89)


def configure_pdf_font(pdf: FPDF) -> str:
    font_path = font_manager.findfont("DejaVu Sans")
    if font_path and os.path.isfile(font_path):
        pdf.add_font("DejaVuSans", "", font_path)
        return "DejaVuSans"
    return "Helvetica"


def make_fv_profile_player_only(report: dict[str, Any]) -> io.BytesIO:
    v0 = float(report["v0"])
    f0 = float(report["f0"])

    x_player = [0, v0]
    y_player = [f0, 0]

    bbox = dict(boxstyle="round", edgecolor="none", facecolor=BLUE_HEX)

    fig, ax = plt.subplots()
    ax.plot(x_player, y_player, label="Player", color=BLUE_HEX, linewidth=2.5)

    ax.annotate(
        str(round(v0, 2)),
        (v0, 0),
        xytext=(v0, -0.6),
        textcoords="data",
        ha="center",
        va="center",
        color="white",
        bbox=bbox,
    )
    ax.annotate(
        str(round(f0, 2)),
        (0, f0),
        xytext=(-0.6, f0),
        textcoords="data",
        ha="center",
        va="center",
        color="white",
        bbox=bbox,
    )

    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.set_xlabel("V0 [m/s]")
    ax.set_ylabel("F0 [N/kg]")
    ax.legend(loc="upper right", frameon=False)
    ax.margins(x=0.05, y=0.05)

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
    buf.seek(0)
    plt.close(fig)
    return buf


def make_norm_scatter_plot(
    report: dict[str, Any],
    norm_row: dict[str, Any],
    scatter_entry: dict[str, Any],
    language: str,
) -> io.BytesIO:
    texts = PDF_TEXT[language]
    points = scatter_entry.get("points") or []
    x_values = [float(point["v0"]) for point in points]
    y_values = [float(point["f0"]) for point in points]
    player_v0 = float(report["v0"])
    player_f0 = float(report["f0"])
    v0_median = float(norm_row["v0_median"])
    f0_median = float(norm_row["f0_median"])

    all_x = [*x_values, player_v0]
    all_y = [*y_values, player_f0]
    x_span = max(all_x) - min(all_x) if all_x else 1.0
    y_span = max(all_y) - min(all_y) if all_y else 1.0
    x_pad = max(x_span * 0.08, 0.15)
    y_pad = max(y_span * 0.08, 0.15)

    fig, ax = plt.subplots()
    if x_values and y_values:
        ax.scatter(x_values, y_values, color="#d0d4db", s=32, alpha=0.9, edgecolors="none")

    ax.scatter([player_v0], [player_f0], color="#FB3331", s=80, zorder=3)
    ax.axhline(f0_median, color="#252423", linestyle="--", linewidth=1.2)
    ax.axvline(v0_median, color="#252423", linestyle="--", linewidth=1.2)

    ax.text(v0_median + x_pad * 0.15, max(all_y) + y_pad * 0.1, "Q1", fontsize=10, color="#252423")
    ax.text(min(all_x) - x_pad * 0.1, max(all_y) + y_pad * 0.1, "Q2", fontsize=10, color="#252423")
    ax.text(min(all_x) - x_pad * 0.1, min(all_y) - y_pad * 0.35, "Q3", fontsize=10, color="#252423")
    ax.text(v0_median + x_pad * 0.15, min(all_y) - y_pad * 0.35, "Q4", fontsize=10, color="#252423")

    ax.set_xlim(min(all_x) - x_pad, max(all_x) + x_pad)
    ax.set_ylim(min(all_y) - y_pad, max(all_y) + y_pad)
    ax.set_xlabel("V0 [m/s]")
    ax.set_ylabel("F0 [N/kg]")
    ax.set_title(f"{texts['quadrant_reference']} | {norm_row.get('category')}")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.grid(True, linestyle="--", alpha=0.18)

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
    buf.seek(0)
    plt.close(fig)
    return buf


def make_normative_fv_profile(
    report: dict[str, Any],
    norm_row: dict[str, Any],
    language: str,
) -> io.BytesIO:
    texts = PDF_TEXT[language]
    player_v0 = float(report["v0"])
    player_f0 = float(report["f0"])
    bbox = dict(boxstyle="round", edgecolor="none", facecolor=BLUE_HEX)
    upper_x = [0, float(norm_row["v0_max"])]
    upper_y = [float(norm_row["f0_max"]), 0]
    lower_x = [0, float(norm_row["v0_min"])]
    lower_y = [float(norm_row["f0_min"]), 0]
    x_span = max(upper_x[1], player_v0) - min(lower_x[1], 0)
    y_span = max(upper_y[0], player_f0) - min(lower_y[0], 0)
    x_pad = max(x_span * 0.08, 0.15)
    y_pad = max(y_span * 0.08, 0.15)

    x_player = [0, player_v0]
    y_player = [player_f0, 0]

    fig, ax_line = plt.subplots(figsize=(5.8, 3.8))
    ax_line.plot(
        upper_x,
        upper_y,
        label=texts["upper_reference"],
        linestyle="--",
        color="#00C060",
        linewidth=2.0,
    )
    ax_line.plot(x_player, y_player, label=texts["player_label"], color=BLUE_HEX, linewidth=2.5)
    ax_line.plot(
        lower_x,
        lower_y,
        label=texts["lower_reference"],
        linestyle="--",
        color="#FB3331",
        linewidth=2.0,
    )
    ax_line.annotate(
        str(round(player_v0, 2)),
        (player_v0, 0),
        xytext=(player_v0, -0.6),
        textcoords="data",
        ha="center",
        va="center",
        color="white",
        bbox=bbox,
    )
    ax_line.annotate(
        str(round(player_f0, 2)),
        (0, player_f0),
        xytext=(-0.6, player_f0),
        textcoords="data",
        ha="center",
        va="center",
        color="white",
        bbox=bbox,
    )
    ax_line.spines["top"].set_visible(False)
    ax_line.spines["right"].set_visible(False)
    ax_line.set_xlabel("V0 [m/s]")
    ax_line.set_ylabel("F0 [N/kg]")
    ax_line.set_title(texts["player_fv_profile"], fontsize=10)
    ax_line.set_xlim(-x_pad * 0.2, max(upper_x[1], player_v0) + x_pad)
    ax_line.set_ylim(-y_pad * 0.2, max(upper_y[0], player_f0) + y_pad)
    ax_line.legend(loc="upper right", frameon=False)
    ax_line.margins(x=0.05, y=0.05)

    buf = io.BytesIO()
    fig.tight_layout()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
    buf.seek(0)
    plt.close(fig)
    return buf


def make_split_speed_time_plot(run: dict[str, Any]) -> io.BytesIO:
    motions = run.get("motions") or []
    fig, ax = plt.subplots(figsize=(5.8, 3.8))

    if motions:
        x_points = [0.0]
        y_points = [float(motions[0].get("avgSpeed") or 0)]
        for motion in motions:
            start_time = float(motion.get("startTime") or 0)
            end_time = float(motion.get("endTime") or 0)
            avg_speed = float(motion.get("avgSpeed") or 0)
            x_points.extend([start_time, end_time])
            y_points.extend([avg_speed, avg_speed])

        ax.plot(x_points, y_points, color=BLUE_HEX, linewidth=2.5)

        for motion in motions:
            mid_time = (float(motion.get("startTime") or 0) + float(motion.get("endTime") or 0)) / 2
            phase_name = str(motion.get("phaseName") or "")
            if phase_name:
                ax.text(
                    mid_time,
                    float(motion.get("avgSpeed") or 0) + 0.15,
                    phase_name,
                    fontsize=8,
                    color="#6b7280",
                    ha="center",
                )

    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.set_xlabel("Čas [s]")
    ax.set_ylabel("Rýchlosť [m/s]")
    ax.set_title("Rýchlosť počas runu", fontsize=10)
    ax.grid(True, linestyle="--", alpha=0.18)

    buf = io.BytesIO()
    fig.tight_layout()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
    buf.seek(0)
    plt.close(fig)
    return buf


def build_non_normative_split_pdf(
    run: dict[str, Any],
    player_name: str,
    logo_bytes: bytes | None,
    player_photo_bytes: bytes | None,
) -> bytes:
    chart_buf = make_split_speed_time_plot(run)
    pdf = FPDF("L", "mm", "A4")
    pdf.set_auto_page_break(auto=False)
    pdf.add_page()
    font_family = configure_pdf_font(pdf)

    left_x = 14
    left_w = 126
    right_x = 154
    right_w = 129
    top_y = 18
    subtitle_y = 29
    chart_y = 54
    chart_w = left_w
    scheme_y = 137
    scheme_w = 96
    scheme_x = left_x + (left_w - scheme_w) / 2
    photo_w = 54
    photo_x = right_x + 4
    photo_y = 16
    logo_w = 30
    logo_x = right_x + right_w - logo_w
    logo_y = 16
    metrics_y = 92
    rs_logo_w = 36
    rs_logo_x = pdf.w - rs_logo_w - 10
    rs_logo_y = pdf.h - 15

    pdf.set_text_color(*BLACK_RGB)
    pdf.set_font(font_family, "", 24)
    pdf.set_xy(left_x, top_y)
    pdf.cell(0, 10, player_name, new_x="LMARGIN", new_y="NEXT")

    pdf.set_text_color(128, 128, 128)
    pdf.set_font(font_family, "", 12)
    pdf.set_xy(left_x, subtitle_y)
    pdf.cell(0, 10, "Deceleračný profil 15-0-5", new_x="LMARGIN", new_y="NEXT")

    pdf.image(chart_buf, x=left_x, y=chart_y, w=chart_w)
    if SPLIT_1505_IMAGE_PATH.is_file():
        pdf.image(str(SPLIT_1505_IMAGE_PATH), x=scheme_x, y=scheme_y, w=scheme_w)

    if player_photo_bytes:
        pdf.image(io.BytesIO(player_photo_bytes), x=photo_x, y=photo_y, w=photo_w, h=photo_w, keep_aspect_ratio=True)
    if logo_bytes:
        pdf.image(io.BytesIO(logo_bytes), x=logo_x, y=logo_y, w=logo_w)

    cell_w = 24
    cell_h = 12
    cell_h_sub = 7
    row_gap = 18
    col_gap = 28
    metrics_block_w = cell_w * 3 + 2 * (col_gap - cell_w)
    data_x = right_x + (right_w - metrics_block_w) / 2

    total_time = float(run.get("time") or 0)
    top_speed_ms = float(run.get("topSpeed") or 0)
    max_acceleration = float(run.get("maxAcceleration") or 0)
    max_deceleration = float(run.get("maxDeceleration") or 0)
    deceleration_time = float(run.get("decelerationTime") or 0)

    pdf.set_fill_color(*BLUE_RGB)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font(font_family, "", 12)

    rounded_corner_cell(pdf, data_x, metrics_y, cell_w, cell_h, format_decimal(total_time))
    rounded_corner_cell(pdf, data_x + col_gap, metrics_y, cell_w, cell_h, format_decimal(top_speed_ms))
    rounded_corner_cell(pdf, data_x + 2 * col_gap, metrics_y, cell_w, cell_h, format_decimal(top_speed_ms * 3.6))
    rounded_corner_cell(pdf, data_x, metrics_y + row_gap, cell_w, cell_h, format_decimal(max_acceleration))
    rounded_corner_cell(pdf, data_x + col_gap, metrics_y + row_gap, cell_w, cell_h, format_decimal(max_deceleration))
    rounded_corner_cell(pdf, data_x + 2 * col_gap, metrics_y + row_gap, cell_w, cell_h, format_decimal(deceleration_time))

    pdf.set_font(font_family, "", 8)
    pdf.set_text_color(220, 220, 220)
    rounded_corner_cell(pdf, data_x, metrics_y + 9, cell_w, cell_h_sub, "Celkový čas [s]")
    rounded_corner_cell(pdf, data_x + col_gap, metrics_y + 9, cell_w, cell_h_sub, "Max rýchlosť [m/s]")
    rounded_corner_cell(pdf, data_x + 2 * col_gap, metrics_y + 9, cell_w, cell_h_sub, "Max rýchlosť [km/h]")
    rounded_corner_cell(pdf, data_x, metrics_y + row_gap + 9, cell_w, cell_h_sub, "Max akcelerácia")
    rounded_corner_cell(pdf, data_x + col_gap, metrics_y + row_gap + 9, cell_w, cell_h_sub, "Max decelerácia")
    rounded_corner_cell(pdf, data_x + 2 * col_gap, metrics_y + row_gap + 9, cell_w, cell_h_sub, "Čas decelerácie")

    if RS_LOGO_PATH.is_file():
        pdf.image(str(RS_LOGO_PATH), x=rs_logo_x, y=rs_logo_y, w=rs_logo_w)

    return bytes(pdf.output(dest="S"))


def make_deceleration_speed_time_plot(run: dict[str, Any], language: str = "Slovak") -> io.BytesIO:
    texts = PDF_TEXT[language]
    samples = run.get("plotSamples") or []
    fig, ax = plt.subplots(figsize=(5.8, 3.8))

    if samples:
        times = [float(sample.get("t_rel") or 0) for sample in samples]
        speeds = [float(sample.get("speed_mps") or 0) for sample in samples]
        accelerations = [float(sample.get("acceleration_mps2") or 0) for sample in samples]
        start_index = int(run.get("startIndex") or 0)
        stop_index = int(run.get("stopIndex") or 0)
        mid_index = int(run.get("midIndex") or start_index)
        v_stop = float(run.get("vStop") or DECEL_V_STOP)

        start_rel_index = max(0, min(start_index, len(samples) - 1))
        stop_rel_index = max(0, min(stop_index, len(samples) - 1))
        mid_rel_index = max(0, min(mid_index, len(samples) - 1))

        segment_times = times[start_rel_index:stop_rel_index + 1]
        segment_speeds = speeds[start_rel_index:stop_rel_index + 1]
        segment_accelerations = accelerations[start_rel_index:stop_rel_index + 1]

        ax.plot(segment_times, segment_speeds, color=BLUE_HEX, linewidth=2.5, zorder=3)

        vmax_index = max(range(len(segment_speeds)), key=lambda index: segment_speeds[index])
        decm_index = min(range(len(segment_accelerations)), key=lambda index: segment_accelerations[index])
        vmax_time = segment_times[vmax_index]
        vmax_speed = segment_speeds[vmax_index]
        decm_time = segment_times[decm_index]
        decm_speed = segment_speeds[decm_index]
        stop_time = segment_times[-1]
        stop_speed = segment_speeds[-1]
        mid_time = times[mid_rel_index]

        ax.scatter(vmax_time, vmax_speed, s=48, color=BLUE_HEX, zorder=5)
        ax.scatter(stop_time, stop_speed, s=48, color=ORANGE_HEX, zorder=5)
        ax.scatter(decm_time, decm_speed, s=58, color=RED_HEX, marker="D", zorder=6)

        ax.annotate("VMax", (vmax_time, vmax_speed), xytext=(6, 6), textcoords="offset points", fontsize=9)
        ax.annotate("Stop", (stop_time, stop_speed), xytext=(6, 6), textcoords="offset points", fontsize=9)
        ax.annotate(
            f"DecM\n{format_decimal(run.get('DecM'))}",
            (decm_time, decm_speed),
            xytext=(8, -14),
            textcoords="offset points",
            fontsize=8,
            ha="left",
            va="top",
        )

        ax.axvspan(segment_times[0], mid_time, color=BLUE_HEX, alpha=0.12, zorder=1)
        ax.axvspan(mid_time, stop_time, color=ORANGE_HEX, alpha=0.10, zorder=1)
        ax.axvline(mid_time, linestyle="--", linewidth=1.1, color="#6b7280", alpha=0.85)
        ax.axvline(stop_time, linestyle="--", linewidth=1.0, color="#6b7280", alpha=0.7)
        ax.axhline(v_stop, linestyle=":", linewidth=1.0, color="#6b7280", alpha=0.6)

        zone_y = max(segment_speeds) * 0.5 if segment_speeds else 0
        ax.text((segment_times[0] + mid_time) / 2, zone_y, texts["early_dec"], ha="center", va="top", fontsize=9, weight="bold")
        ax.text((mid_time + stop_time) / 2, zone_y, texts["late_dec"], ha="center", va="top", fontsize=9, weight="bold")

        ax.annotate(
            "",
            xy=(stop_time, stop_speed),
            xytext=(segment_times[0], stop_speed),
            arrowprops=dict(arrowstyle="<->", lw=1.3, alpha=0.9, color="#252423"),
        )
        ax.text(
            stop_time / 2,
            stop_speed + (max(segment_speeds) - min(segment_speeds)) * 0.08,
            f"TTS = {format_decimal(run.get('TTS'))} s",
            ha="center",
            va="bottom",
            fontsize=9,
            weight="bold",
        )

    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.set_xlabel("Time [s]" if language == "English" else "Čas [s]")
    ax.set_ylabel("Speed [m/s]" if language == "English" else "Rýchlosť [m/s]")
    ax.set_title(texts["decel_chart_title"], fontsize=10)
    ax.grid(True, linestyle="--", alpha=0.18)
    ax.margins(x=0.02, y=0.20)

    buf = io.BytesIO()
    fig.tight_layout()
    fig.savefig(buf, format="png", dpi=150, bbox_inches="tight", transparent=True)
    buf.seek(0)
    plt.close(fig)
    return buf


def build_non_normative_deceleration_pdf(
    run: dict[str, Any],
    player_name: str,
    exercise_name: str,
    logo_bytes: bytes | None,
    player_photo_bytes: bytes | None,
    language: str,
) -> bytes:
    texts = PDF_TEXT[language]
    chart_buf = make_deceleration_speed_time_plot(run, language)
    pdf = FPDF("L", "mm", "A4")
    pdf.set_auto_page_break(auto=False)
    pdf.add_page()
    font_family = configure_pdf_font(pdf)

    left_x = 14
    left_w = 126
    right_x = 154
    right_w = 129
    top_y = 18
    subtitle_y = 29
    exercise_y = 39
    chart_y = 56
    chart_w = left_w
    photo_w = 46
    photo_x = right_x + (right_w - photo_w) / 2
    photo_y = 14
    logo_w = 32
    logo_x = right_x + right_w - logo_w - 4
    logo_y = 14
    metrics_y = 86
    rs_logo_w = 36
    rs_logo_x = pdf.w - rs_logo_w - 10
    rs_logo_y = pdf.h - 15

    if logo_bytes:
        pdf.image(io.BytesIO(logo_bytes), x=logo_x, y=logo_y, w=logo_w)

    pdf.set_text_color(*BLACK_RGB)
    pdf.set_font(font_family, "", 24)
    pdf.set_xy(left_x, top_y)
    pdf.cell(0, 10, player_name, new_x="LMARGIN", new_y="NEXT")

    pdf.set_text_color(128, 128, 128)
    pdf.set_font(font_family, "", 12)
    pdf.set_xy(left_x, subtitle_y)
    pdf.cell(0, 10, texts["decel_title"], new_x="LMARGIN", new_y="NEXT")

    pdf.set_font(font_family, "", 11)
    pdf.set_xy(left_x, exercise_y)
    pdf.cell(0, 10, exercise_name, new_x="LMARGIN", new_y="NEXT")

    pdf.image(chart_buf, x=left_x, y=chart_y, w=chart_w)

    if player_photo_bytes:
        pdf.image(io.BytesIO(player_photo_bytes), x=photo_x, y=photo_y, w=photo_w, h=photo_w, keep_aspect_ratio=True)

    cell_w = 24
    cell_h = 12
    cell_h_sub = 7
    row_gap = 18
    col_gap = 28
    metrics_block_w = cell_w * 3 + 2 * (col_gap - cell_w)
    data_x = right_x + (right_w - metrics_block_w) / 2

    pdf.set_fill_color(*BLUE_RGB)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font(font_family, "", 12)

    rounded_corner_cell(pdf, data_x, metrics_y, cell_w, cell_h, format_decimal(run.get("averageDeceleration")))
    rounded_corner_cell(pdf, data_x + col_gap, metrics_y, cell_w, cell_h, format_decimal(run.get("DecM")))
    rounded_corner_cell(pdf, data_x + 2 * col_gap, metrics_y, cell_w, cell_h, format_decimal(run.get("VMax")))
    rounded_corner_cell(pdf, data_x, metrics_y + row_gap, cell_w, cell_h, format_decimal(run.get("TTS")))
    rounded_corner_cell(pdf, data_x + col_gap, metrics_y + row_gap, cell_w, cell_h, format_decimal(run.get("DTS")))

    pdf.set_font(font_family, "", 8)
    pdf.set_text_color(220, 220, 220)
    rounded_corner_cell(pdf, data_x, metrics_y + 9, cell_w, cell_h_sub, texts["decel_avg"])
    rounded_corner_cell(pdf, data_x + col_gap, metrics_y + 9, cell_w, cell_h_sub, texts["decel_max"])
    rounded_corner_cell(pdf, data_x + 2 * col_gap, metrics_y + 9, cell_w, cell_h_sub, texts["decel_vmax"])
    rounded_corner_cell(pdf, data_x, metrics_y + row_gap + 9, cell_w, cell_h_sub, texts["decel_tts"])
    rounded_corner_cell(pdf, data_x + col_gap, metrics_y + row_gap + 9, cell_w, cell_h_sub, texts["decel_dts"])

    if RS_LOGO_PATH.is_file():
        pdf.image(str(RS_LOGO_PATH), x=rs_logo_x, y=rs_logo_y, w=rs_logo_w)

    return bytes(pdf.output(dest="S"))


def build_non_normative_fv_pdf(
    report: dict[str, Any],
    player_name: str,
    logo_bytes: bytes | None,
    player_photo_bytes: bytes | None,
    language: str,
) -> bytes:
    texts = PDF_TEXT[language]
    fv_buf = make_fv_profile_player_only(report)
    pdf = FPDF("L", "mm", "A4")
    pdf.set_auto_page_break(auto=False)
    pdf.add_page()
    font_family = configure_pdf_font(pdf)

    if logo_bytes:
        pdf.image(io.BytesIO(logo_bytes), x=10, y=12, w=25)

    pdf.set_text_color(*BLACK_RGB)
    pdf.set_font(font_family, "", 32)
    pdf.set_xy(45, 18)
    pdf.cell(0, 10, player_name, new_x="LMARGIN", new_y="NEXT")

    pdf.set_text_color(128, 128, 128)
    pdf.set_font(font_family, "", 14)
    pdf.set_xy(45, 34)
    pdf.cell(0, 10, texts["fv_title"], new_x="LMARGIN", new_y="NEXT")

    pdf.image(fv_buf, x=20, y=55, w=140)
    if player_photo_bytes:
        pdf.image(io.BytesIO(player_photo_bytes), x=194, y=18, w=45, h=45, keep_aspect_ratio=True)

    data_x = 170
    data_y = 70
    y_second_row = 18
    cell_w = 28
    cell_h = 12
    cell_h_sub = 7

    pdf.set_fill_color(*BLUE_RGB)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font(font_family, "", 12)

    f0 = float(report["f0"])
    v0 = float(report["v0"])
    pmax = float(report["pMax"])
    drf = float(report.get("ratioOfForceDecrease") or 0)
    rfmax = float(report.get("ratioOfForceMax") or 0)

    rounded_corner_cell(pdf, data_x, data_y, cell_w, cell_h, str(round(f0, 2)))
    rounded_corner_cell(pdf, data_x + 32, data_y, cell_w, cell_h, str(round(v0, 2)))
    rounded_corner_cell(pdf, data_x + 64, data_y, cell_w, cell_h, str(round(pmax, 2)))
    rounded_corner_cell(pdf, data_x, data_y + y_second_row, cell_w, cell_h, str(round(v0 * 3.6, 2)))
    rounded_corner_cell(pdf, data_x + 32, data_y + y_second_row, cell_w, cell_h, str(round(drf, 2)))
    rounded_corner_cell(pdf, data_x + 64, data_y + y_second_row, cell_w, cell_h, str(round(rfmax, 2)))

    pdf.set_font(font_family, "", 8)
    pdf.set_text_color(220, 220, 220)
    rounded_corner_cell(pdf, data_x, data_y + 9, cell_w, cell_h_sub, "F0 [N/kg]")
    rounded_corner_cell(pdf, data_x + 32, data_y + 9, cell_w, cell_h_sub, "V0 [m/s]")
    rounded_corner_cell(pdf, data_x + 64, data_y + 9, cell_w, cell_h_sub, "PMax [W]")
    rounded_corner_cell(pdf, data_x, data_y + 9 + y_second_row, cell_w, cell_h_sub, "V0 [km/h]")
    rounded_corner_cell(pdf, data_x + 32, data_y + 9 + y_second_row, cell_w, cell_h_sub, "DRF")
    rounded_corner_cell(pdf, data_x + 64, data_y + 9 + y_second_row, cell_w, cell_h_sub, "RFmax")

    return bytes(pdf.output(dest="S"))


def load_fv_norms() -> tuple[list[dict[str, Any]], str | None]:
    if not FV_NORMS_PATH.is_file():
        return [], f"Norms file not found: {FV_NORMS_PATH}"

    try:
        workbook = load_workbook(FV_NORMS_PATH, data_only=True)
    except Exception as exc:
        return [], f"Could not read norms file: {exc}"

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], "The norms workbook is empty."

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    required = ["category", "f0_min", "f0_max", "v0_min", "v0_max", "f0_median", "v0_median"]
    missing = [name for name in required if name not in headers]
    if missing:
        return [], f"The norms workbook is missing columns: {', '.join(missing)}"

    items: list[dict[str, Any]] = []
    for row in rows[1:]:
        item = {headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))}
        if item.get("category"):
            items.append(item)

    return items, None


def load_fv_norm_scatter() -> tuple[dict[str, dict[str, Any]], str | None]:
    if not FV_NORM_SCATTER_PATH.is_file():
        return {}, f"Norm scatter file not found: {FV_NORM_SCATTER_PATH}"

    try:
        payload = json.loads(FV_NORM_SCATTER_PATH.read_text(encoding="utf-8"))
    except Exception as exc:
        return {}, f"Could not read norm scatter file: {exc}"

    if not isinstance(payload, dict):
        return {}, "The norm scatter file did not return the expected object format."

    return payload, None


def get_norm_quadrant(report: dict[str, Any], norm_row: dict[str, Any]) -> str:
    f0 = float(report["f0"])
    v0 = float(report["v0"])
    f0_mid = float(norm_row["f0_median"])
    v0_mid = float(norm_row["v0_median"])

    if v0 > v0_mid and f0 > f0_mid:
        return "Q1"
    if v0 <= v0_mid and f0 > f0_mid:
        return "Q2"
    if v0 <= v0_mid and f0 <= f0_mid:
        return "Q3"
    return "Q4"


def get_quadrant_result(quadrant: str, language: str) -> str:
    texts = PDF_TEXT[language]
    return texts[f"{quadrant.lower()}_result"]


def get_quadrant_recommendations(quadrant: str, language: str) -> list[str]:
    texts = PDF_TEXT[language]
    return texts[f"{quadrant.lower()}_recs"]


def build_normative_fv_pdf(
    report: dict[str, Any],
    player_name: str,
    logo_bytes: bytes | None,
    player_photo_bytes: bytes | None,
    norm_row: dict[str, Any],
    scatter_entry: dict[str, Any] | None,
    language: str,
) -> bytes:
    texts = PDF_TEXT[language]
    fv_buf = make_normative_fv_profile(report, norm_row, language)
    scatter_buf = (
        make_norm_scatter_plot(report, norm_row, scatter_entry, language)
        if scatter_entry
        else None
    )
    pdf = FPDF("L", "mm", "A4")
    pdf.set_auto_page_break(auto=False)
    pdf.add_page()
    font_family = configure_pdf_font(pdf)

    left_x = 14
    left_w = 126
    right_x = 154
    right_w = 129
    logo_w = 32
    top_y = 18
    subtitle_y = 29
    badge_y = 39
    header_bottom_y = badge_y + 11
    logo_x = pdf.w - logo_w - 12
    logo_y = top_y + ((header_bottom_y - top_y) - logo_w) / 2
    has_logo = bool(logo_bytes)
    identity_x = left_x
    fv_chart_y = 56
    fv_chart_w = left_w
    metrics_y = 148
    rec_title_y = 154
    rec_text_y = 161
    photo_w = 46
    photo_x = right_x + (right_w - photo_w) / 2
    photo_y = 14
    scatter_y = 66
    scatter_w = 96
    scatter_x = right_x + (right_w - scatter_w) / 2
    rs_logo_w = 36
    rs_logo_x = pdf.w - rs_logo_w - 10
    rs_logo_y = pdf.h - 15

    if logo_bytes:
        pdf.image(io.BytesIO(logo_bytes), x=logo_x, y=logo_y, w=logo_w)

    pdf.set_text_color(*BLACK_RGB)
    pdf.set_font(font_family, "", 24)
    pdf.set_xy(identity_x, top_y)
    pdf.cell(0, 10, player_name, new_x="LMARGIN", new_y="NEXT")

    pdf.set_text_color(128, 128, 128)
    pdf.set_font(font_family, "", 12)
    pdf.set_xy(identity_x, subtitle_y)
    pdf.cell(0, 10, texts["fv_title"], new_x="LMARGIN", new_y="NEXT")

    quadrant = get_norm_quadrant(report, norm_row)
    quadrant_result = get_quadrant_result(quadrant, language)
    recommendations = get_quadrant_recommendations(quadrant, language)

    badge_x = identity_x
    pdf.set_font(font_family, "", 12)
    pdf.set_fill_color(*get_quadrant_badge_fill(quadrant))
    pdf.set_text_color(255, 255, 255)
    rounded_corner_cell(pdf, badge_x, badge_y, 14, 11, quadrant)

    pdf.set_fill_color(*get_quadrant_result_fill(quadrant))
    rounded_corner_cell(pdf, badge_x + 16, badge_y, min(100, left_w - 16), 11, quadrant_result)

    pdf.image(fv_buf, x=left_x, y=fv_chart_y, w=fv_chart_w)
    if player_photo_bytes:
        pdf.image(io.BytesIO(player_photo_bytes), x=photo_x, y=photo_y, w=photo_w, h=photo_w, keep_aspect_ratio=True)
    if scatter_buf:
        pdf.image(scatter_buf, x=scatter_x, y=scatter_y, w=scatter_w)

    f0 = float(report["f0"])
    v0 = float(report["v0"])
    pmax = float(report["pMax"])
    drf = float(report.get("ratioOfForceDecrease") or 0)
    rfmax = float(report.get("ratioOfForceMax") or 0)

    metrics_block_w = 74
    data_x = left_x + (left_w - metrics_block_w) / 2
    data_y = metrics_y
    y_second_row = 18
    cell_w = 22
    cell_h = 12
    cell_h_sub = 7

    pdf.set_fill_color(*BLUE_RGB)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font(font_family, "", 12)

    rounded_corner_cell(pdf, data_x, data_y, cell_w, cell_h, str(round(f0, 2)))
    rounded_corner_cell(pdf, data_x + 26, data_y, cell_w, cell_h, str(round(v0, 2)))
    rounded_corner_cell(pdf, data_x + 52, data_y, cell_w, cell_h, str(round(pmax, 2)))
    rounded_corner_cell(pdf, data_x, data_y + y_second_row, cell_w, cell_h, str(round(v0 * 3.6, 2)))
    rounded_corner_cell(pdf, data_x + 26, data_y + y_second_row, cell_w, cell_h, str(round(drf, 2)))
    rounded_corner_cell(pdf, data_x + 52, data_y + y_second_row, cell_w, cell_h, str(round(rfmax, 2)))

    pdf.set_font(font_family, "", 8)
    pdf.set_text_color(220, 220, 220)
    rounded_corner_cell(pdf, data_x, data_y + 9, cell_w, cell_h_sub, "F0 [N/kg]")
    rounded_corner_cell(pdf, data_x + 26, data_y + 9, cell_w, cell_h_sub, "V0 [m/s]")
    rounded_corner_cell(pdf, data_x + 52, data_y + 9, cell_w, cell_h_sub, "PMax [W]")
    rounded_corner_cell(pdf, data_x, data_y + 9 + y_second_row, cell_w, cell_h_sub, "V0 [km/h]")
    rounded_corner_cell(pdf, data_x + 26, data_y + 9 + y_second_row, cell_w, cell_h_sub, "DRF")
    rounded_corner_cell(pdf, data_x + 52, data_y + 9 + y_second_row, cell_w, cell_h_sub, "RFmax")

    pdf.set_font(font_family, "", 14)
    pdf.set_text_color(*BLACK_RGB)
    pdf.text(right_x, rec_title_y, texts["recommendation"])
    pdf.set_font(font_family, "", 8)
    y_coordinate = rec_text_y
    for item in recommendations:
        pdf.text(right_x, y_coordinate, "* " + item)
        y_coordinate += 7

    if RS_LOGO_PATH.is_file():
        pdf.image(str(RS_LOGO_PATH), x=rs_logo_x, y=rs_logo_y, w=rs_logo_w)

    return bytes(pdf.output(dest="S"))


def load_exercise_report(
    api_key: str,
    exercise_id: str,
    report_type: str,
    exercise: dict[str, Any] | None = None,
) -> bool:
    cache_key = f"{report_type}:{exercise_id}"

    if report_type == "fv":
        payload, error = fetch_force_velocity_exercise(api_key, exercise_id)
    else:
        payload, error = load_split_payload_for_exercise(api_key, exercise_id, exercise)

    if error:
        st.session_state["exercise_report_errors"][cache_key] = error
        st.session_state["exercise_report_cache"].pop(cache_key, None)
        return False

    st.session_state["exercise_report_cache"][cache_key] = payload
    st.session_state["exercise_report_errors"].pop(cache_key, None)
    return True


@st.dialog("FV PDF export")
def render_fv_export_dialog(
    exercise: dict[str, Any],
    reports: list[dict[str, Any]],
    runner_info: dict[str, Any] | None,
    client: dict[str, Any],
) -> None:
    run_options = {
        f"Run {index + 1} | F0 {format_decimal(report.get('f0'))} | V0 {format_decimal(report.get('v0'))}": report
        for index, report in enumerate(reports)
    }

    st.caption("Choose the run, report style, and optional assets for the export.")
    selected_run_label = st.selectbox(
        "Run",
        options=list(run_options.keys()),
        key=f"fv_run_select_{exercise.get('id')}",
    )
    selected_report = run_options[selected_run_label]

    export_col1, export_col2 = st.columns([1, 1])
    export_mode = export_col1.radio(
        "Report type",
        options=["Non-normative", "Normative"],
        horizontal=True,
        key=f"fv_export_mode_{exercise.get('id')}",
    )
    export_language = st.selectbox(
        "Language",
        options=["English", "Slovak"],
        key=f"fv_export_language_{exercise.get('id')}",
    )
    selected_norm = None
    selected_norm_category = ""
    scatter_entry = None
    if export_mode == "Normative":
        norms, norms_error = load_fv_norms()
        scatter_map, scatter_error = load_fv_norm_scatter()
        if norms_error:
            st.error(norms_error)
            return
        if scatter_error:
            st.warning(scatter_error)
            scatter_map = {}

        norm_options = {str(item.get("category")): item for item in norms}
        selected_norm_category = st.selectbox(
            "Reference category",
            options=list(norm_options.keys()),
            key=f"fv_norm_category_{exercise.get('id')}",
        )
        selected_norm = norm_options[selected_norm_category]
        scatter_entry = scatter_map.get(selected_norm_category)

    logo_bytes = render_logo_library_selector("shared_logo", export_language)

    selected_runner = selected_report.get("runnerInfo") or runner_info or {}
    export_name = (
        format_optional_value(selected_runner.get("displayName"))
        if format_optional_value(selected_runner.get("displayName")) != "-"
        else format_optional_value(client.get("displayName"))
    )
    player_client_id = str(
        selected_runner.get("clientId")
        or client.get("id")
        or ""
    )

    player_photo_bytes = render_player_photo_selector(
        f"shared_player_photo_{player_client_id}",
        player_client_id,
        export_language,
    ) if player_client_id else None

    if export_mode == "Normative" and selected_norm is not None:
        missing_values = [
            key
            for key in ("f0_median", "v0_median")
            if selected_norm.get(key) in (None, "")
        ]
        if missing_values:
            st.warning(
                "Selected norm category is missing values in the workbook: "
                + ", ".join(missing_values)
            )
            return

        pdf_bytes = fv_build_normative_fv_pdf(
            selected_report,
            export_name,
            logo_bytes,
            player_photo_bytes,
            selected_norm,
            scatter_entry,
            PDF_TEXT[export_language],
        )
        file_name = (
            f"{safe_filename(export_name)}_{safe_filename(selected_norm_category)}_"
            f"{safe_filename(str(exercise.get('id')))}_fv_normative.pdf"
        )
        st.download_button(
            "Download normative FV PDF",
            data=pdf_bytes,
            file_name=file_name,
            mime="application/pdf",
            key=f"fv_pdf_normative_{exercise.get('id')}",
            width="stretch",
        )
    else:
        pdf_bytes = fv_build_non_normative_fv_pdf(
            selected_report,
            export_name,
            logo_bytes,
            player_photo_bytes,
            PDF_TEXT[export_language],
        )
        file_name = f"{safe_filename(export_name)}_{safe_filename(str(exercise.get('id')))}_fv_profile.pdf"
        st.download_button(
            "Download FV PDF",
            data=pdf_bytes,
            file_name=file_name,
            mime="application/pdf",
            key=f"fv_pdf_download_{exercise.get('id')}",
            width="stretch",
        )


@st.dialog("15-0-5 PDF export")
def render_split_export_dialog(
    exercise: dict[str, Any],
    payload: dict[str, Any],
    client: dict[str, Any],
) -> None:
    derived_runs = payload.get("_derived_runs") or []
    if not derived_runs:
        st.info("No 15-0-5 runs are available for PDF export.")
        return

    run_options = {
        f"Run {index + 1} | Time {format_decimal(run.get('time'), 3)} | Top speed {format_decimal(run.get('topSpeed'))}": run
        for index, run in enumerate(derived_runs)
    }

    st.caption("Choose the run and optional assets for the export.")
    selected_run_label = st.selectbox(
        "Run",
        options=list(run_options.keys()),
        key=f"split_run_select_{exercise.get('id')}",
    )
    selected_run = run_options[selected_run_label]

    export_language = st.selectbox(
        "Language",
        options=["English", "Slovak"],
        key=f"split_export_language_{exercise.get('id')}",
    )

    logo_bytes = render_logo_library_selector("shared_logo", export_language)
    export_name = format_optional_value(client.get("displayName"))
    player_client_id = str(client.get("id") or "")
    player_photo_bytes = render_player_photo_selector(
        f"shared_player_photo_{player_client_id}",
        player_client_id,
        export_language,
    ) if player_client_id else None

    pdf_bytes = split_build_non_normative_split_pdf(
        selected_run,
        export_name,
        logo_bytes,
        player_photo_bytes,
        export_language,
    )
    file_name = f"{safe_filename(export_name)}_{safe_filename(str(exercise.get('id')))}_1505_profile.pdf"
    st.download_button(
        "Download 15-0-5 PDF",
        data=pdf_bytes,
        file_name=file_name,
        mime="application/pdf",
        key=f"split_pdf_download_{exercise.get('id')}",
        width="stretch",
    )


@st.dialog("Deceleration PDF export")
def render_deceleration_export_dialog(
    exercise: dict[str, Any],
    payload: dict[str, Any],
    client: dict[str, Any],
) -> None:
    export_language = st.selectbox(
        PDF_TEXT["English"]["language"],
        options=["English", "Slovak"],
        format_func=lambda option: PDF_TEXT[option]["english"] if option == "English" else PDF_TEXT[option]["slovak"],
        key=f"decel_export_language_{exercise.get('id')}",
    )
    texts = PDF_TEXT[export_language]
    deceleration_runs = payload.get("_deceleration_runs") or []
    if not deceleration_runs:
        st.info(texts["decel_no_runs"])
        return

    run_options = {
        (
            f"Run {index + 1} | DecA {format_decimal(run.get('averageDeceleration'))} | "
            f"VMax {format_decimal(run.get('VMax'))}"
        ): run
        for index, run in enumerate(deceleration_runs)
    }

    st.caption(texts["decel_choose_run"])
    selected_run_label = st.selectbox(
        texts["decel_run"],
        options=list(run_options.keys()),
        key=f"decel_run_select_{exercise.get('id')}",
    )
    selected_run = run_options[selected_run_label]

    logo_bytes = render_logo_library_selector("shared_logo", export_language)
    export_name = format_optional_value(client.get("displayName"))
    player_client_id = str(client.get("id") or "")
    player_photo_bytes = render_player_photo_selector(
        f"shared_player_photo_{player_client_id}",
        player_client_id,
        export_language,
    ) if player_client_id else None

    exercise_name = str(
        selected_run.get("exerciseName")
        or exercise.get("name")
        or exercise.get("exerciseTypeName")
        or "Deceleration"
    )
    pdf_bytes = decel_build_non_normative_deceleration_pdf(
        selected_run,
        export_name,
        exercise_name,
        logo_bytes,
        player_photo_bytes,
        texts,
        export_language,
    )
    file_name = f"{safe_filename(export_name)}_{safe_filename(str(exercise.get('id')))}_deceleration_profile.pdf"
    st.download_button(
        texts["decel_download_pdf"],
        data=pdf_bytes,
        file_name=file_name,
        mime="application/pdf",
        key=f"decel_pdf_download_{exercise.get('id')}",
        width="stretch",
    )


def render_fv_profile(exercise: dict[str, Any], payload: dict[str, Any], client: dict[str, Any]) -> None:
    reports = payload.get("reports") or []
    failed_reports = payload.get("failedReports") or []
    runner_info = reports[0].get("runnerInfo") if reports else None

    st.markdown("### Running (LR) force-velocity profile")
    st.caption(f"Valid runs: {len(reports)} | Failed runs: {len(failed_reports)}")

    if reports:
        if st.button(
            "Open FV PDF export",
            key=f"fv_export_open_top_{exercise.get('id')}",
            width="stretch",
        ):
            render_fv_export_dialog(exercise, reports, runner_info, client)

        summary_rows = [
            {
                "Run": index + 1,
                "F0": format_decimal(report.get("f0")),
                "V0": format_decimal(report.get("v0")),
                "PMax": format_decimal(report.get("pMax")),
                "Tau": format_decimal(report.get("tau")),
                "Confidence": format_decimal(report.get("confidence"), 3),
                "Est. unloaded max speed": format_decimal(report.get("estimatedUnloadedMaxSpeed")),
                "RF max": format_decimal(report.get("ratioOfForceMax")),
                "Force decrease": format_decimal(report.get("ratioOfForceDecrease")),
            }
            for index, report in enumerate(reports)
        ]
        st.dataframe(summary_rows, width="stretch", hide_index=True)
        st.caption("Review the runs, then use the export button above when you are ready to create a PDF.")
    else:
        st.info("No FV runs were returned for this exercise.")


def render_split_profile(exercise: dict[str, Any], payload: dict[str, Any], client: dict[str, Any]) -> None:
    reports = payload.get("reports") or []
    derived_runs = payload.get("_derived_runs") or []

    st.markdown("### 15-0-5 split profile")

    top_col1, top_col2, top_col3 = st.columns(3)
    displayed_run_count = len(reports) if reports else len(derived_runs)
    top_col1.metric("Runs", displayed_run_count)
    top_col2.metric("Split length", format_decimal(reports[0].get("splitLength")) if reports else "5.00")
    top_col3.metric("Units", "meters" if (reports or derived_runs) else "-")

    if derived_runs:
        if st.button(
            "Open 15-0-5 PDF export",
            key=f"split_export_open_top_{exercise.get('id')}",
            width="stretch",
        ):
            render_split_export_dialog(exercise, payload, client)

    split_rows: list[dict[str, Any]] = []
    for report_index, report in enumerate(reports):
        for split in report.get("splits") or []:
            split_rows.append(
                {
                    "Run": report_index + 1,
                    "Start": format_decimal(split.get("start")),
                    "End": format_decimal(split.get("end")),
                    "Time": format_decimal(split.get("time"), 3),
                    "Top speed": format_decimal(split.get("topSpeed")),
                    "Max force": format_decimal(split.get("maxForce")),
                    "Max power": format_decimal(split.get("maxPower")),
                }
            )

    if split_rows:
        st.dataframe(split_rows, width="stretch", hide_index=True)
    elif derived_runs:
        derived_rows = [
            {
                "Run": index + 1,
                "Distance": format_decimal(run.get("distance")),
                "Time": format_decimal(run.get("time"), 3),
                "Top speed": format_decimal(run.get("topSpeed")),
                "Load": format_decimal(run.get("load")),
            }
            for index, run in enumerate(derived_runs)
        ]
        st.dataframe(derived_rows, width="stretch", hide_index=True)
    else:
        st.info("No split rows were returned for this exercise.")


def render_deceleration_profile(
    exercise: dict[str, Any],
    payload: dict[str, Any],
    client: dict[str, Any],
) -> None:
    deceleration_runs = payload.get("_deceleration_runs") or []
    texts = PDF_TEXT["English"]

    st.markdown(f"### {texts['decel_title']}")

    top_col1, top_col2, top_col3 = st.columns(3)
    top_col1.metric(texts["decel_runs"], len(deceleration_runs))
    top_col2.metric(texts["decel_threshold"], format_decimal(DECEL_ACC_THRESHOLD))
    top_col3.metric(texts["decel_stop_speed"], format_decimal(DECEL_V_STOP))

    if deceleration_runs:
        if st.button(
            texts["decel_open_export"],
            key=f"decel_export_open_top_{exercise.get('id')}",
            width="stretch",
        ):
            render_deceleration_export_dialog(exercise, payload, client)

        preview_options = {
            (
                f"Run {index + 1} | DecA {format_decimal(run.get('averageDeceleration'))} | "
                f"VMax {format_decimal(run.get('VMax'))}"
            ): run
            for index, run in enumerate(deceleration_runs)
        }
        selected_preview_label = st.selectbox(
            texts["decel_preview_run"],
            options=list(preview_options.keys()),
            key=f"decel_preview_{exercise.get('id')}",
        )
        selected_run = preview_options[selected_preview_label]

        chart_col, info_col = st.columns([1.45, 1], vertical_alignment="top")
        with chart_col:
            st.image(
                decel_make_deceleration_speed_time_plot(selected_run, texts, "English"),
                use_container_width=True,
            )
        with info_col:
            exercise_name = str(
                selected_run.get("exerciseName")
                or exercise.get("name")
                or exercise.get("exerciseTypeName")
                or "-"
            )
            st.caption(exercise_name)
            st.dataframe(
                [
                    {
                        texts["decel_avg"]: format_decimal(selected_run.get("averageDeceleration")),
                        texts["decel_max"]: format_decimal(selected_run.get("DecM")),
                        texts["decel_vmax"]: format_decimal(selected_run.get("VMax")),
                        texts["decel_tts"]: format_decimal(selected_run.get("TTS")),
                        texts["decel_dts"]: format_decimal(selected_run.get("DTS")),
                    }
                ],
                width="stretch",
                hide_index=True,
            )
    else:
        st.info(texts["decel_no_valid_runs"])


def render_deceleration_debug(
    exercise: dict[str, Any],
    payload: dict[str, Any] | None,
) -> None:
    with st.expander(f"Debug deceleration fetch: {exercise.get('exerciseTypeName') or exercise.get('name') or exercise.get('id')}"):
        debug_info = {
            "exerciseId": exercise.get("id"),
            "exerciseName": exercise.get("name"),
            "exerciseTypeName": exercise.get("exerciseTypeName"),
            "setIds": [str(exercise_set.get("id") or "") for exercise_set in (exercise.get("sets") or [])],
            "hasPayload": bool(payload),
            "reportsCount": len((payload or {}).get("reports") or []),
            "derivedRunsCount": len((payload or {}).get("_derived_runs") or []),
            "decelerationRunsCount": len((payload or {}).get("_deceleration_runs") or []),
            "decelerationRunDebug": [
                {
                    "motionGroupId": run.get("motionGroupId"),
                    "plotSampleCount": len(run.get("plotSamples") or []),
                    "isFallback": bool(run.get("isFallback")),
                    "debug": run.get("_debug") or {},
                }
                for run in ((payload or {}).get("_deceleration_runs") or [])
            ],
            "decelerationRawFailures": (payload or {}).get("_deceleration_raw_failures") or [],
            "debugFetches": (payload or {}).get("_debug_fetches") or [],
            "decelerationRuns": (payload or {}).get("_deceleration_runs") or [],
        }
        st.code(json.dumps(sanitize_for_debug(debug_info), ensure_ascii=False, indent=2), language="json")


def render_session_detail_content(
    session_detail: dict[str, Any],
    client_lookup: dict[str, dict[str, Any]],
) -> None:
    exercises = session_detail.get("exercises") or []
    client = client_lookup.get(str(session_detail.get("clientId") or ""), {})
    session_time = format_session_timestamp(session_detail.get("timestamp"))

    st.subheader("Selected session")
    st.caption("Session overview and exercise-specific reports.")

    if exercises:
        exercise_rows = [
            {
                "Exercise": exercise.get("exerciseTypeName"),
                "Sets": len(exercise.get("sets") or []),
            }
            for exercise in exercises
        ]
        st.dataframe(exercise_rows, width="stretch", hide_index=True)

        api_key = st.session_state.get("api_key", "")
        fv_exercises = [
            exercise
            for exercise in exercises
            if str(exercise.get("exerciseTypeName") or "").strip() == "Running (LR)"
        ]
        split_exercises = [
            exercise
            for exercise in exercises
            if str(exercise.get("exerciseTypeName") or "").strip() == "15-0-5"
        ]
        deceleration_exercises = [
            exercise
            for exercise in exercises
            if str(exercise.get("exerciseTypeName") or "").strip() != "Running (LR)"
        ]

        for exercise in fv_exercises:
            cache_key = f"fv:{exercise.get('id')}"
            if cache_key not in st.session_state["exercise_report_cache"] and api_key:
                with st.spinner("Loading FV profile data..."):
                    load_exercise_report(api_key, str(exercise.get("id") or ""), "fv")

            error = st.session_state["exercise_report_errors"].get(cache_key)
            if error:
                st.error(error)
            else:
                payload = st.session_state["exercise_report_cache"].get(cache_key)
                if payload:
                    render_fv_profile(exercise, payload, client)

        for exercise in split_exercises:
            cache_key = f"split:{exercise.get('id')}"
            if cache_key not in st.session_state["exercise_report_cache"] and api_key:
                with st.spinner("Loading deceleration split data..."):
                    load_exercise_report(api_key, str(exercise.get("id") or ""), "split", exercise)

            error = st.session_state["exercise_report_errors"].get(cache_key)
            if error:
                st.error(error)
            else:
                payload = st.session_state["exercise_report_cache"].get(cache_key)
                if (
                    payload
                    and not (payload.get("reports") or [])
                    and "_debug_fetches" not in payload
                    and api_key
                ):
                    with st.spinner("Refreshing split debug data..."):
                        load_exercise_report(api_key, str(exercise.get("id") or ""), "split", exercise)
                    payload = st.session_state["exercise_report_cache"].get(cache_key)
                if payload:
                    render_split_profile(exercise, payload, client)

        for exercise in deceleration_exercises:
            cache_key = f"split:{exercise.get('id')}"
            if cache_key not in st.session_state["exercise_report_cache"] and api_key:
                with st.spinner("Loading deceleration profile data..."):
                    load_exercise_report(api_key, str(exercise.get("id") or ""), "split", exercise)

            error = st.session_state["exercise_report_errors"].get(cache_key)
            if error:
                continue

            payload = st.session_state["exercise_report_cache"].get(cache_key)
            if (
                payload
                and "_deceleration_runs" not in payload
                and api_key
            ):
                with st.spinner("Refreshing deceleration debug data..."):
                    load_exercise_report(api_key, str(exercise.get("id") or ""), "split", exercise)
                payload = st.session_state["exercise_report_cache"].get(cache_key)

            render_deceleration_debug(exercise, payload)
            if payload and (payload.get("_deceleration_runs") or []):
                render_deceleration_profile(exercise, payload, client)
    else:
        st.info("No exercises were returned for this session.")


def render_session_selection_block(
    sessions: list[dict[str, Any]],
    client_lookup: dict[str, dict[str, Any]],
    scope: str,
    table_key: str,
) -> None:
    ensure_session_scope_defaults(scope)
    selected_session_id = st.session_state.get(f"selected_session_id_{scope}", "")

    for index, session in enumerate(sessions):
        session_id = str(session.get("id") or "")
        session_time = format_session_timestamp(session.get("timestamp"))
        athlete_name = get_client_display_name(session.get("clientId"), client_lookup)
        is_selected = session_id == selected_session_id

        with st.container(border=True):
            row_col1, row_col2 = st.columns([5, 1.1], vertical_alignment="center")
            with row_col1:
                st.markdown(
                    f"""
                    <div class="session-row-inline">
                      <p class="session-row-title">{session_time}</p>
                      <p class="session-row-subtitle">{athlete_name}</p>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )
            button_label = "Selected" if is_selected else "Open session"
            button_type = "secondary" if is_selected else "primary"
            if row_col2.button(
                button_label,
                key=f"{table_key}_open_{index}",
                width="stretch",
                type=button_type,
            ) and not is_selected:
                st.session_state[f"selected_session_id_{scope}"] = session_id
                with st.spinner("Loading session detail..."):
                    load_selected_session_detail(scope)
                st.rerun()

    session_detail_error = st.session_state.get(f"session_detail_error_{scope}", "")
    if session_detail_error:
        st.error(session_detail_error)

    session_detail = st.session_state.get(f"session_detail_{scope}")
    if session_detail:
        render_session_detail_content(session_detail, client_lookup)


def render_client_detail(client: dict[str, Any]) -> None:
    st.subheader("Athlete overview")
    st.caption("Profile details and recent sessions for the selected athlete.")

    top_col1, top_col2, top_col3 = st.columns(3)
    top_col1.metric("Name", format_optional_value(client.get("displayName")))
    top_col2.metric("Group", format_optional_value(client.get("group")))
    top_col3.metric("Date of birth", format_optional_value(client.get("dateOfBirth")))

    info_col1, info_col2 = st.columns(2)
    with info_col1:
        st.write(f"**Height:** {format_optional_value(client.get('height'))}")
        st.write(f"**Weight:** {format_optional_value(client.get('weight'))}")
    with info_col2:
        st.write(f"**Created:** {format_optional_value(client.get('created'))}")
        st.write(f"**Edited:** {format_optional_value(client.get('edited'))}")
        st.write(f"**Tags:** {format_optional_value(client.get('tags'))}")

    st.subheader("Sessions")

    filter_col1, filter_col2, filter_col3 = st.columns([1, 1, 0.8])
    filter_col1.date_input("From", key="session_filter_from")
    filter_col2.date_input("To", key="session_filter_to")
    reload_sessions = filter_col3.button("Refresh sessions", width="stretch")
    filters_changed = (
        st.session_state["session_filter_from"] != st.session_state.get("session_filter_last_from")
        or st.session_state["session_filter_to"] != st.session_state.get("session_filter_last_to")
    )

    if st.session_state["session_filter_from"] > st.session_state["session_filter_to"]:
        st.error("The start date must be earlier than or equal to the end date.")
        return

    if reload_sessions or filters_changed:
        st.session_state["session_filter_last_from"] = st.session_state["session_filter_from"]
        st.session_state["session_filter_last_to"] = st.session_state["session_filter_to"]
        with st.spinner("Loading sessions for the selected client..."):
            load_sessions_for_selected_client()

    if st.session_state["client_sessions_error"]:
        st.error(st.session_state["client_sessions_error"])

    sessions = st.session_state.get("client_sessions", [])
    st.caption(
        f"Showing {len(sessions)} sessions from "
        f"{st.session_state['session_filter_from'].isoformat()} to {st.session_state['session_filter_to'].isoformat()}"
    )

    if sessions:
        render_session_selection_block(
            sessions=sessions,
            client_lookup=get_client_lookup(st.session_state.get("clients_cache", [])),
            scope="client",
            table_key="client_sessions_table",
        )
    else:
        st.info("No sessions found for the selected range.")

def render_dashboard() -> None:
    api_key = st.session_state["api_key"]
    storage_key = storage_namespace(api_key)
    recent_opened_storage_key = recent_opened_sessions_namespace(api_key)
    auth_user_email = st.session_state.get("auth_user_email", "")

    storage_result = CLIENT_STORAGE_COMPONENT(
        data={
            "storageKey": storage_key,
            "command": st.session_state["client_storage_command"],
            "commandId": st.session_state["client_storage_command_id"],
            "clientsJson": st.session_state["client_storage_payload"],
            "lastSynced": st.session_state["clients_last_synced"],
            "recentOpenedStorageKey": recent_opened_storage_key,
            "recentOpenedCommand": st.session_state["recent_opened_command"],
            "recentOpenedCommandId": st.session_state["recent_opened_command_id"],
            "recentOpenedSessionsJson": st.session_state["recent_opened_sessions_payload"],
        },
        default={"clients_json": "", "last_synced": "", "recent_opened_sessions_json": ""},
        on_clients_json_change=lambda: None,
        on_last_synced_change=lambda: None,
        on_recent_opened_sessions_json_change=lambda: None,
        key="client_storage_bridge",
        height=0,
    )

    cached_clients_json = storage_result.clients_json or ""
    cached_last_synced = storage_result.last_synced or ""
    cached_recent_opened_json = getattr(storage_result, "recent_opened_sessions_json", "") or ""

    if (
        not st.session_state["client_storage_autoload_complete"]
        and cached_clients_json
        and not st.session_state["clients_cache"]
    ):
        try:
            st.session_state["clients_cache"] = json.loads(cached_clients_json)
            st.session_state["clients_last_synced"] = cached_last_synced
            st.session_state["client_storage_command"] = "read"
            st.session_state["client_storage_payload"] = cached_clients_json
            st.session_state["client_storage_autoload_complete"] = True
            st.rerun()
        except json.JSONDecodeError:
            st.session_state["client_storage_error"] = "Client cache in local storage is corrupted."
            st.session_state["client_storage_autoload_complete"] = True

    if (
        not st.session_state["client_storage_autoload_complete"]
        and not cached_clients_json
        and not st.session_state["clients_cache"]
    ):
        with st.spinner("Loading clients from the 1080 API..."):
            loaded = load_clients_from_api(api_key)
        st.session_state["client_storage_autoload_complete"] = True
        if loaded:
            st.rerun()

    if not st.session_state["recent_opened_autoload_complete"]:
        if cached_recent_opened_json:
            try:
                cached_recent_opened = json.loads(cached_recent_opened_json)
                pruned_recent_opened = prune_recent_opened_sessions(cached_recent_opened)
                st.session_state["recent_opened_sessions"] = pruned_recent_opened
                st.session_state["recent_opened_autoload_complete"] = True
                if json.dumps(pruned_recent_opened) != cached_recent_opened_json:
                    sync_recent_opened_sessions_to_storage(pruned_recent_opened)
                    st.rerun()
            except json.JSONDecodeError:
                st.session_state["recent_opened_sessions"] = []
                st.session_state["recent_opened_autoload_complete"] = True
        else:
            st.session_state["recent_opened_sessions"] = []
            st.session_state["recent_opened_autoload_complete"] = True

    st.markdown(
        """
        <div class="hero-card">
          <h1 style="margin:0;">1080 Reports</h1>
          <p class="section-intro">Browse athletes, review sessions, and export clean performance reports.</p>
        </div>
        """,
        unsafe_allow_html=True,
    )
    if auth_user_email:
        st.caption(f"Signed in as {auth_user_email}")

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Status", "Connected")
    col2.metric("API", "1080 Motion")
    col3.metric("Clients", len(st.session_state["clients_cache"]))
    col4.metric("Last synced", format_sync_label(st.session_state["clients_last_synced"]))

    action_col1, action_col2 = st.columns([1, 1])
    if action_col1.button("Reload clients", width="stretch"):
        with st.spinner("Reloading clients from the 1080 API..."):
            loaded = load_clients_from_api(api_key)
        st.session_state["client_storage_autoload_complete"] = True
        if loaded:
            st.rerun()

    action_col2.button("Sign out", on_click=logout, width="stretch")

    if st.session_state["client_storage_error"]:
        st.error(st.session_state["client_storage_error"])

    overview_tab, athletes_tab = st.tabs(["Recent sessions", "Athletes"])

    with overview_tab:
        st.markdown("### Recent sessions")
        st.caption("A quick view of the latest activity from the last 7 days.")
        recent_from = date.today() - timedelta(days=7)
        recent_to = date.today()
        with st.spinner("Loading sessions from the last 7 days..."):
            recent_sessions, recent_sessions_error = fetch_recent_sessions(api_key, recent_from, recent_to)

        if recent_sessions_error:
            st.error(recent_sessions_error)
        else:
            client_lookup = get_client_lookup(st.session_state.get("clients_cache", []))
            if recent_sessions and has_missing_client_names(recent_sessions, client_lookup):
                with st.spinner("Refreshing clients to resolve missing session names..."):
                    loaded = load_clients_from_api(api_key)
                if loaded:
                    st.rerun()

            st.caption(
                f"Showing {len(recent_sessions or [])} sessions from {recent_from.isoformat()} to {recent_to.isoformat()}"
            )
            if recent_sessions:
                render_session_selection_block(
                    sessions=recent_sessions,
                    client_lookup=get_client_lookup(st.session_state.get("clients_cache", [])),
                    scope="recent",
                    table_key="recent_sessions_table",
                )
            else:
                st.info("No sessions were found in the last 7 days.")

            recent_opened_sessions = st.session_state.get("recent_opened_sessions", [])
            if recent_opened_sessions:
                st.markdown("### Recent opened sessions")
                st.caption("Recently opened from the Athletes section. Stored for 1 day.")
                render_session_selection_block(
                    sessions=recent_opened_sessions,
                    client_lookup=get_client_lookup(st.session_state.get("clients_cache", [])),
                    scope="recent_opened",
                    table_key="recent_opened_sessions_table",
                )

    with athletes_tab:
        st.markdown("### Athletes")
        st.caption("Search and filter the athlete list, then open a profile to review session history.")

        clients = st.session_state["clients_cache"]
        if clients:
            group_options = sorted(
                {
                    str(client.get("group")).strip()
                    for client in clients
                    if client.get("group") not in (None, "")
                }
            )

            filter_col1, filter_col2 = st.columns([2, 1])
            search_query = filter_col1.text_input(
                "Search athletes",
                placeholder="Search by athlete name or external reference",
            ).strip()
            selected_group = filter_col2.selectbox(
                "Group",
                options=["All groups", *group_options],
                index=0,
            )

            normalized_query = search_query.lower()
            filtered_clients = []
            for client in clients:
                name = str(client.get("displayName") or "")
                external_id = str(client.get("externalId") or "")
                group = str(client.get("group") or "")

                matches_group = selected_group == "All groups" or group == selected_group
                matches_query = (
                    not normalized_query
                    or normalized_query in name.lower()
                    or normalized_query in external_id.lower()
                )

                if matches_group and matches_query:
                    filtered_clients.append(client)

            filtered_clients = sorted(
                filtered_clients,
                key=lambda client: str(client.get("displayName") or "").lower(),
            )

            visible_clients = filtered_clients if normalized_query else filtered_clients[:10]

            preview = [
                {
                    "Name": client.get("displayName"),
                    "Group": client.get("group"),
                    "Tags": format_optional_value(client.get("tags")),
                }
                for client in visible_clients
            ]

            if normalized_query:
                st.caption(f"Showing {len(visible_clients)} matching athletes")
            else:
                st.caption(f"Showing first {len(visible_clients)} athletes alphabetically out of {len(filtered_clients)}")

            selected_client_id = st.session_state.get("selected_client_id", "")
            table_event = st.dataframe(
                preview,
                width="stretch",
                hide_index=True,
                on_select="rerun",
                selection_mode="single-row",
                key="clients_table",
            )

            selected_rows = table_event.selection.rows
            if selected_rows:
                selected_client = visible_clients[selected_rows[0]]
                st.session_state["selected_client_id"] = str(selected_client.get("id") or "")

            selected_client_id = st.session_state.get("selected_client_id", "")
            selected_client = next(
                (
                    client
                    for client in clients
                    if str(client.get("id") or "") == selected_client_id
                ),
                None,
            )

            if selected_client:
                if st.session_state["selected_client_last_id"] != selected_client_id:
                    st.session_state["selected_client_last_id"] = selected_client_id
                    st.session_state["session_filter_from"] = date.today() - timedelta(days=7)
                    st.session_state["session_filter_to"] = date.today()
                    st.session_state["session_filter_last_from"] = st.session_state["session_filter_from"]
                    st.session_state["session_filter_last_to"] = st.session_state["session_filter_to"]
                    st.session_state["client_sessions"] = []
                    st.session_state["client_sessions_error"] = ""
                    with st.spinner("Loading sessions for the selected client..."):
                        load_sessions_for_selected_client()
                render_client_detail(selected_client)
        else:
            st.info("No clients are currently cached.")


def main() -> None:
    st.set_page_config(
        page_title="1080 Reports",
        page_icon="🚀",
        layout="wide",
    )

    render_app_styles()
    ensure_session_defaults()
    handle_auth_storage_bridge()

    if not st.session_state["auth_verified"] and not st.session_state["auth_storage_autoload_complete"]:
        st.caption("Checking saved session...")
        return

    if st.session_state["auth_verified"] and st.session_state["api_valid"]:
        render_dashboard()
    else:
        render_login()


if __name__ == "__main__":
    main()
